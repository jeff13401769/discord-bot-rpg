import datetime
import openpyxl
from openpyxl.styles import Alignment
import pytz
import time
import functools
import os
import difflib

import discord
from discord import Option, OptionChoice
from discord.ext import commands, tasks

from utility.config import config
from cogs.function_in import function_in
from cogs.function_in_in import function_in_in
from cogs.aibot import Aibot

class Info(discord.Cog, name="資訊"):
    def __init__(self, bot):
        self.bot: discord.Bot = bot
        
    async def players_autocomplete(self, ctx: discord.AutocompleteContext):
        query = ctx.value.lower() if ctx.value else ""
        
        members = await function_in.sql_findall('rpg_players', 'players')
        members_list = []
        for member in members:
            user = self.bot.get_user(member[0])
            if not user:
                name = f"機器人無法獲取名稱 ({member[0]})"
            else:
                name = f"{user.name} ({user.id})"
            members_list.append(name)
        
        if query:
            # 依相似度排序，越接近輸入的越前面
            members_list = sorted(
                members_list,
                key=lambda x: difflib.SequenceMatcher(None, query, x.lower()).ratio(),
                reverse=True
            )
            members_list = [m for m in members_list if query in m.lower() or difflib.SequenceMatcher(None, query, m.lower()).ratio() > 0.3]
        return members_list[:25]
        
    @discord.user_command(name="rpg資訊", description="查看自己或別人的資訊",
        options=[
            discord.Option(
                discord.Member,
                name="玩家",
                description="選擇欲查看的玩家",
                required=False
            )
        ]
    )
    async def rpg資訊(self, interaction: discord.ApplicationContext, players: discord.Member):
        await self.資訊(interaction, f"{players.name} ({players.id})")

    @commands.slash_command(name="資訊", description="查看自己或別人的資訊",
        options=[
            discord.Option(
                str,
                name="玩家",
                description="選擇欲查看的玩家",
                required=False,
                autocomplete=players_autocomplete
            )
        ]
    )
    async def 資訊(self, interaction: discord.ApplicationContext, players: str):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        if players:
            players = await function_in.players_list_to_players(self, players)
            checkreg = await function_in.checkreg(self, interaction, players.id)
            if not checkreg:
                return
            player_id = players.id
            func_user = players
        else:
            player_id = user.id
            func_user = user
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, player_id)
        embed = discord.Embed(title=f"{func_user.name} 的屬性介面", color=0xFF0000)
        if func_user.avatar:
            embed.set_thumbnail(url=f"{func_user.avatar.url}")
        else:
            embed.set_thumbnail(url=f"{func_user.default_avatar.url}")
        embed.add_field(name="玩家:", value=f"{func_user.mention}", inline=True)
        special_exp = 1
        check_special = await function_in.check_special(self, func_user.id, players_class)
        if check_special:
            special_exp = 2
        #if players_level < 12:
        #    expfull = int(19.5 * 1.95 ** players_level) * special_exp
        #else:
        expfull = int((17 * players_level) ** 1.7) * special_exp
        exp_100_no = (players_exp / expfull) * 100
        exp_100 = round(exp_100_no)
        embed.add_field(name="職業:", value=f"{players_class}", inline=True)
        embed.add_field(name="等級:", value=f"{players_level} ({players_exp}/{expfull}) ({exp_100}%)", inline=True)
        embed.add_field(name="飽食度:", value=f"{players_hunger}🍗", inline=False)
        embed.add_field(name="血量:", value=f"{players_hp}/{players_max_hp}💖", inline=True)
        embed.add_field(name="魔力:", value=f"{players_mana}/{players_max_mana}💧", inline=True)
        embed.add_field(name="物理攻擊力:", value=f"{players_AD}⚔", inline=True)
        embed.add_field(name="魔法攻擊力:", value=f"{players_AP}🔮", inline=True)
        embed.add_field(name="防禦力:", value=f"{players_def}🛡", inline=True)
        embed.add_field(name="爆擊率:", value=f"{players_crit_chance}%💥", inline=True)
        embed.add_field(name="爆擊傷害:", value=f"{players_crit_damage}%🧨", inline=True)
        embed.add_field(name="閃避率:", value=f"{players_dodge}%🌟", inline=True)
        embed.add_field(name="命中率:", value=f"{players_hit}%✨", inline=True)
        embed.add_field(name="破甲率:", value=f"{players_ndef}%✨", inline=True)
        embed.add_field(name="\u200b", value="\u200b", inline=False)
        embed.add_field(name="力量:", value=f"{players_str} <:str:1087788396447010956>", inline=True)
        embed.add_field(name="智慧:", value=f"{players_int} <:int:1087789657569382522>", inline=True)
        embed.add_field(name="敏捷:", value=f"{players_dex} <:dex:1087792723609788517>", inline=True)
        embed.add_field(name="體質:", value=f"{players_con} <:con:1087794018911522997>", inline=True)
        embed.add_field(name="幸運:", value=f"{players_luk} <:luk:1087794455760883784>", inline=True)
        embed.add_field(name="\u200b", value="\u200b", inline=False)
        embed.add_field(name="屬性點:", value=f"{players_attr_point+players_add_attr_point}", inline=True)
        embed.add_field(name="天賦點:", value=f"{players_skill_point}", inline=True)
        embed.add_field(name="\u200b", value="\u200b", inline=False)
        players_register_time = datetime.datetime.fromtimestamp(players_register_time)
        embed.add_field(name="註冊時間:", value=f"{players_register_time}", inline=False)
        await interaction.followup.send(embed=embed, view=self.info_menu(interaction, func_user))

    class info_menu(discord.ui.View):
        def __init__(self, interaction: discord.ApplicationContext, player: discord.Member):
            super().__init__(timeout=30)
            self.interaction = interaction
            self.player = player
            self.button1 = discord.ui.Button(emoji="📘", label="基礎", style=discord.ButtonStyle.blurple, custom_id="button1")
            self.button2 = discord.ui.Button(emoji="<:equipment:1078600684624171068>", label="裝備", style=discord.ButtonStyle.blurple, custom_id="button2")
            self.button3 = discord.ui.Button(emoji="<:medal:1146472732494659615>", label="勳章", style=discord.ButtonStyle.blurple, custom_id="button3")
            self.button4 = discord.ui.Button(emoji="<:mage:975947436885430362>", label="技能", style=discord.ButtonStyle.blurple, custom_id="button4")
            self.button5 = discord.ui.Button(emoji="<:coin:1078582446091665438>", label="貨幣", style=discord.ButtonStyle.blurple, custom_id="button5")
            self.button6 = discord.ui.Button(emoji="🍽️", label="料理", style=discord.ButtonStyle.blurple, custom_id="button6")
            self.button7 = discord.ui.Button(emoji="<:buff:1398047025118969988>", label="Buff", style=discord.ButtonStyle.blurple, custom_id="button7")
            self.button8 = discord.ui.Button(emoji="<a:sword:1219469485875138570>", label="PVP面板", style=discord.ButtonStyle.blurple, custom_id="button8")
            self.button9 = discord.ui.Button(emoji="🤖", label="小幫手", style=discord.ButtonStyle.blurple, custom_id="button9")
            self.button10 = discord.ui.Button(emoji="📰", label="股票", style=discord.ButtonStyle.blurple, custom_id="button10")
            self.button1.callback = functools.partial(self.button1_callback, interaction)
            self.button2.callback = functools.partial(self.button2_callback, interaction)
            self.button3.callback = functools.partial(self.button3_callback, interaction)
            self.button4.callback = functools.partial(self.button4_callback, interaction)
            self.button5.callback = functools.partial(self.button5_callback, interaction)
            self.button6.callback = functools.partial(self.button6_callback, interaction)
            self.button7.callback = functools.partial(self.button7_callback, interaction)
            self.button8.callback = functools.partial(self.button8_callback, interaction)
            self.button9.callback = functools.partial(self.button9_callback, interaction)
            self.button10.callback = functools.partial(self.button10_callback, interaction)
            self.add_item(self.button1)
            self.add_item(self.button2)
            self.add_item(self.button3)
            self.add_item(self.button4)
            self.add_item(self.button5)
            self.add_item(self.button6)
            self.add_item(self.button7)
            self.add_item(self.button8)
            self.add_item(self.button9)
            self.add_item(self.button10)

        async def on_timeout(self):
            await super().on_timeout()
            self.disable_all_items()
            if self.interaction.message:
                try:
                    await self.interaction.message.edit(view=None)
                    self.stop()
                except discord.errors.InteractionResponded:
                    pass
            else:
                self.stop()

        async def button1_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
            embed = discord.Embed(title=f"{user.name} 的屬性介面", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=True)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            special_exp = 1
            check_special = await function_in.check_special(self, user.id, players_class)
            if check_special:
                special_exp = 2
            #if players_level < 12:
            #    expfull = int(19.5 * 1.95 ** players_level) * special_exp
            #else:
            expfull = int((17 * players_level) ** 1.7) * special_exp
            exp_100_no = (players_exp / expfull) * 100
            exp_100 = round(exp_100_no)
            embed.add_field(name="職業:", value=f"{players_class}", inline=True)
            embed.add_field(name="等級:", value=f"{players_level} ({players_exp}/{expfull}) ({exp_100}%)", inline=True)
            embed.add_field(name="飽食度:", value=f"{players_hunger}🍗", inline=False)
            embed.add_field(name="血量:", value=f"{players_hp}/{players_max_hp}💖", inline=True)
            embed.add_field(name="魔力:", value=f"{players_mana}/{players_max_mana}💧", inline=True)
            embed.add_field(name="物理攻擊力:", value=f"{players_AD}⚔", inline=True)
            embed.add_field(name="魔法攻擊力:", value=f"{players_AP}🔮", inline=True)
            embed.add_field(name="防禦力:", value=f"{players_def}🛡", inline=True)
            embed.add_field(name="爆擊率:", value=f"{players_crit_chance}%💥", inline=True)
            embed.add_field(name="爆擊傷害:", value=f"{players_crit_damage}%🧨", inline=True)
            embed.add_field(name="閃避率:", value=f"{players_dodge}%🌟", inline=True)
            embed.add_field(name="命中率:", value=f"{players_hit}%✨", inline=True)
            embed.add_field(name="破甲率:", value=f"{players_ndef}%✨", inline=True)
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            embed.add_field(name="力量:", value=f"{players_str} <:str:1087788396447010956>", inline=True)
            embed.add_field(name="智慧:", value=f"{players_int} <:int:1087789657569382522>", inline=True)
            embed.add_field(name="敏捷:", value=f"{players_dex} <:dex:1087792723609788517>", inline=True)
            embed.add_field(name="體質:", value=f"{players_con} <:con:1087794018911522997>", inline=True)
            embed.add_field(name="幸運:", value=f"{players_luk} <:luk:1087794455760883784>", inline=True)
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            embed.add_field(name="屬性點:", value=f"{players_attr_point+players_add_attr_point}", inline=True)
            embed.add_field(name="天賦點:", value=f"{players_skill_point}", inline=True)
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            players_register_time = datetime.datetime.fromtimestamp(players_register_time)
            embed.add_field(name="註冊時間:", value=f"{players_register_time}", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()

        async def button2_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的裝備介面", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=False)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            item_type_list = ["武器","頭盔","胸甲","護腿","鞋子","副手","戒指","項鍊","披風","護身符","卡牌欄位1","卡牌欄位2","卡牌欄位3", "職業專用道具"]
            for item_type in item_type_list:
                search = await function_in.sql_search("rpg_equip", f"{user.id}", ["slot"], [item_type])
                equip = search[1]
                embed.add_field(name=f"{item_type}: {equip}", value="\u200b", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()

        async def button3_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的勳章介面", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=False)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            players_medal_list = await function_in.sql_search("rpg_players", "players", ["user_id"], [user.id])
            medal_list = players_medal_list[17]
            if medal_list == "":
                embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            else:
                medal_list = medal_list.split(",")
                for medal in medal_list:
                    if medal != "":
                        embed.add_field(name=f"{medal}", value="\u200b", inline=True)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()

        async def button4_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的技能介面", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=False)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            che = False
            skills = await function_in.sql_findall("rpg_skills", f"{user.id}")
            for skill in skills:
                embed.add_field(name=f"{skill[0]}", value=f"等級: {skill[1]} | 技能熟練度: {skill[2]}", inline=False)
                che=True
            if not che:
                embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            if len(embed.fields) > 25:
                del embed.fields[24:]
                embed.add_field(name="由於超過Discord Embed 25行限制, 以下已被省略...", value="已將您的技能表以Excel發送至您的私聊", inline=False)
                workbook = openpyxl.Workbook()
                alignment = Alignment(horizontal='center', vertical='center')
                sheet1 = workbook.active
                sheet1.title = '技能表'
                sheet2 = workbook.create_sheet(title='生成時間')
                sheets = workbook.sheetnames
                sheet1['A1'] = '技能名稱'
                sheet1['B1'] = '技能等級'
                sheet1['C1'] = '技能熟練度'
                sheet2['A1'] = '本工作表生成時間'
                a = 0
                for skill in skills:
                    sheet1[f'A{a+1}'] = f'{skill[0]}'
                    sheet1[f'B{a+1}'] = f'{skill[1]}'
                    sheet1[f'C{a+1}'] = f'{skill[2]}'
                    a+=1
                now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime("%Y年%m月%d日-%H:%M:%S")
                sheet2['B1'] = f"{now_time}"
                for sheet_name in sheets:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.alignment = alignment
                    sheet.column_dimensions['A'].width = 50
                    sheet.column_dimensions['B'].width = 10
                    sheet.column_dimensions['C'].width = 10
                    if sheet.title == "生成時間":
                        sheet.column_dimensions['A'].width = 30
                        sheet.column_dimensions['B'].width = 100
                base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
                folder_path = os.path.join(base_path, "excel_cache")
                save_filename = f"{user.id}_skills.xlsx"
                save_path = os.path.join(folder_path, save_filename)
                workbook.save(save_path)
                workbook.close()
                file = discord.File(save_path)
                await user.send(file=file)
                os.remove(save_path)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()

        async def button5_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
            embed = discord.Embed(title=f"{user.name} 的貨幣介面", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=True)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            embed.add_field(name="晶幣:", value=f"{players_money}<:coin:1078582446091665438>", inline=False)
            embed.add_field(name="水晶:", value=f"{players_diamond}💎", inline=False)
            embed.add_field(name="任務點數:", value=f"{players_qp}🧿", inline=False)
            embed.add_field(name="世界幣:", value=f"{players_wbp}<:king:1154993624765956156>", inline=False)
            embed.add_field(name="決鬥點數:", value=f"{players_pp}<a:sword:1219469485875138570>", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()

        async def button6_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的料理介面", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=True)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            
            now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
            timeString = now_time
            struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
            time_stamp = int(time.mktime(struct_time))
            players_food_check = await function_in.sql_check_table("rpg_food", f"{user.id}")
            if players_food_check:
                players_food_list = await function_in.sql_findall("rpg_food", f"{user.id}")
                if players_food_list:
                    for food_info in players_food_list:
                        food = food_info[0]
                        food_time_stamp = food_info[1]
                        food_time = await function_in_in.time_calculate(food_time_stamp - time_stamp)
                        embed.add_field(name=f"{food}", value=f"剩餘時間: {food_time}", inline=False)
                else:
                    embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            else:
                embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            if len(embed.fields) > 24:
                del embed.fields[24:]
                embed.add_field(name="由於超過Discord Embed 25行限制, 以下已被省略...", value="...", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()

        async def button7_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的Buff介面", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=True)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            
            now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
            timeString = now_time
            struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
            time_stamp = int(time.mktime(struct_time))
            players_buff_check = await function_in.sql_check_table("rpg_buff", f"{user.id}")
            if players_buff_check:
                players_buff_list = await function_in.sql_findall("rpg_buff", f"{user.id}")
                if players_buff_list:
                    for buff_info in players_buff_list:
                        buff = buff_info[0]
                        buff_time_stamp = buff_info[1]
                        buff_time = await function_in_in.time_calculate(buff_time_stamp - time_stamp)
                        embed.add_field(name=f"{buff}", value=f"剩餘時間: {buff_time}", inline=False)
                else:
                    embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            else:
                embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            if len(embed.fields) > 24:
                del embed.fields[24:]
                embed.add_field(name="由於超過Discord Embed 25行限制, 以下已被省略...", value="...", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()

        async def button8_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的PVP面板", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=True)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            
            players_equip_max_hp = 0
            players_equip_max_mana = 0
            players_equip_def = 0
            players_equip_AD = 0
            players_equip_AP = 0
            players_equip_crit_damage = 0
            players_equip_crit_chance = 0
            players_equip_dodge = 0
            players_equip_ndef = 0
            players_equip_hit = 0
            players_equip_str = 0
            players_equip_int = 0
            players_equip_dex = 0
            players_equip_con = 0
            players_equip_luk = 0

            equips = await function_in.sql_findall("rpg_equip", f"{user.id}")
            set_effects = {}
            for equip in equips:
                slot = equip[0]
                equip = equip[1]
                if "戰鬥道具" in slot or "技能" in slot:
                    continue
                if "無" in equip or "未解鎖" in equip:
                    continue
                data = await function_in.search_for_file(self, equip)
                for attname, value in data.get(equip).get("增加屬性", {}).items():
                    if attname == "PVP_增加血量上限":
                        players_equip_max_hp += value
                    elif attname == "PVP_增加魔力上限":
                        players_equip_max_mana += value
                    elif attname == "PVP_物理攻擊力":
                        players_equip_AD += value
                    elif attname == "PVP_魔法攻擊力":
                        players_equip_AP += value
                    elif attname == "PVP_防禦力":
                        players_equip_def += value
                    elif attname == "PVP_爆擊率":
                        players_equip_crit_chance += value
                    elif attname == "PVP_爆擊傷害":
                        players_equip_crit_damage += value
                    elif attname == "PVP_閃避率":
                        players_equip_dodge += value
                    elif attname == "PVP_命中率":
                        players_equip_hit += value
                    elif attname == "PVP_破甲率":
                        players_equip_ndef += value
                    elif attname == "PVP_力量":
                        players_equip_str += value
                    elif attname == "PVP_智慧":
                        players_equip_int += value
                    elif attname == "PVP_敏捷":
                        players_equip_dex += value
                    elif attname == "PVP_體質":
                        players_equip_con += value
                    elif attname == "PVP_幸運":
                        players_equip_luk += value
                        
                    elif "套裝" in attname:
                        if attname in set_effects:
                            set_effects[attname] += 1
                        else:
                            set_effects[attname] = 1

            if set_effects:
                for set_effect, set_effect_num in set_effects.items():
                    if set_effect == "PVP銅牌套裝":
                        if set_effect_num >= 2:
                            players_equip_dodge += 20
                            players_equip_def += 20
                            players_equip_crit_chance += 20
                        if set_effect_num >= 3:
                            players_equip_dodge += 20
                            players_equip_def += 20
                            players_equip_crit_chance += 20
                            players_equip_crit_damage += 20
                        if set_effect_num >= 4:
                            players_equip_dodge += 20
                            players_equip_def += 20
                            players_equip_crit_chance += 20
                            players_equip_crit_damage += 20
                            players_equip_max_hp += 300

            embed.add_field(name="PVP中裝備增加屬性:", value="\u200b", inline=False)
            attributes = {
                "增加血量上限:": players_equip_max_hp,
                "增加魔力上限:": players_equip_max_mana,
                "物理攻擊力:": players_equip_AD,
                "魔法攻擊力:": players_equip_AP,
                "防禦力:": players_equip_def,
                "爆擊率:": players_equip_crit_chance,
                "爆擊傷害:": players_equip_crit_damage,
                "閃避率:": players_equip_dodge,
                "命中率:": players_equip_hit,
                "破甲率:": players_equip_ndef,
                "力量:": players_equip_str,
                "智慧:": players_equip_int,
                "敏捷:": players_equip_dex,
                "體質:": players_equip_con,
                "幸運:": players_equip_luk,
            }
            for name, value in attributes.items():
                embed.add_field(name=name, value=f"{value}", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
            self.stop()
        
        async def button9_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的小幫手", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=False)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            search = await function_in.sql_search("rpg_system", "daily", ["user_id"], [user.id])
            candaily = search[1]
            if candaily:
                embed.add_field(name=":date: 簽到狀態:", value=":x: 今日尚未簽到", inline=False)
            else:
                embed.add_field(name=":date: 簽到狀態:", value=":white_check_mark: 今日已簽到", inline=False)
            search = await function_in.sql_search("rpg_system", "month_card", ["user_id"], [user.id])
            if search:
                day = search[1]
                embed.add_field(name=":credit_card: 星辰之約狀態:", value=f":white_check_mark: 目前月卡剩餘 {day} 天", inline=False)
            else:
                embed.add_field(name=":credit_card: 星辰之約狀態:", value=":x: 當前無有效的月卡", inline=False)
            now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime("%Y-%m-%d %H:%M:%S")
            timeString = now_time
            struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
            time_stamp = int(time.mktime(struct_time))
            search = await function_in.sql_search("rpg_players", "players", ["user_id"], [user.id])
            action = search[14]
            actiontime = await function_in_in.time_calculate(action-time_stamp)
            if time_stamp <= action:
                embed.add_field(name="<a:clock:1220799040782864405> 冷卻條:", value=f":x: 你還需要等待 {actiontime}!", inline=False)
            else:
                embed.add_field(name="<a:clock:1220799040782864405> 冷卻條:", value=":white_check_mark: 可以進行行動!", inline=False)
            search = await function_in.sql_search("rpg_players", "players", ["user_id"], [user.id])
            players_level = search[1]
            players_all_attr_point = search[20]
            if int(players_level*0.1)*5 < players_all_attr_point:
                embed.add_field(name="<:rpg_boost:1382689893129388073> 神性之石:", value=f":x: 當前已使用 {players_all_attr_point} 顆神性之石, 當前已無法使用更多神性之石", inline=False)
            else:
                embed.add_field(name="<:rpg_boost:1382689893129388073> 神性之石:", value=f":white_check_mark: 當前已使用 {players_all_attr_point} 顆神性之石, 還可以使用 {int(players_level*0.1)*5 - players_all_attr_point} 顆神性之石", inline=False)
            affection, amount = await Aibot.check_favorability(self, user)
            if not affection:
                embed.add_field(name="<:ehh:1381359837476032612> 拜神:", value=f":x: 於50等時開放", inline=False)
            else:
                if amount > 0:
                    embed.add_field(name="<:ehh:1381359837476032612> 拜神:", value=f":white_check_mark: 本日還可以拜訪兔神 - 雪月‧緋綾 {amount} 次", inline=False)
                else:
                    embed.add_field(name="<:ehh:1381359837476032612> 拜神:", value=f":x: 本日已經拜訪兔神 - 雪月‧緋綾 太多次了", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
        
        async def button10_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = self.player
            embed = discord.Embed(title=f"{user.name} 的股票", color=0xFF0000)
            embed.add_field(name="玩家:", value=f"{user.mention}", inline=False)
            if user.avatar:
                embed.set_thumbnail(url=f"{user.avatar.url}")
            else:
                embed.set_thumbnail(url=f"{user.default_avatar.url}")
            
            stock_check = await function_in.sql_check_table("rpg_stock", f"{user.id}")
            if stock_check:
                stock_list = await function_in.sql_findall("rpg_stock", user.id)
                if stock_list:
                    for stock in stock_list:
                        stock_id = stock[0]
                        stock_amount = stock[1]
                        check = await function_in.sql_search("rpg_stock", "all", ["stock_id"], [stock_id])
                        if not check:
                            await function_in.sql_delete("rpg_stock", user.id, "stock_id", stock_id)
                            continue
                        stock_name = check[1]
                        embed.add_field(name=f"股票名稱: {stock_name}", value=f"股票數量: {stock_amount} 張", inline=False)
                else:
                    embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            else:
                await function_in.sql_create_table("rpg_stock", user.id, ["stock_id", "amount"], ["BIGINT", "BIGINT"], "stock_id")
                embed.add_field(name="空空如也.....", value="\u200b", inline=False)
            if len(embed.fields) > 24:
                del embed.fields[24:]
                embed.add_field(name="由於超過Discord Embed 25行限制, 以下已被省略...", value="...", inline=False)
            await msg.edit(view=Info.info_menu(interaction, user), embed=embed)
                

def setup(client: discord.Bot):
    client.add_cog(Info(client))
