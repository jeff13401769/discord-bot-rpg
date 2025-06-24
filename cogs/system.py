import datetime
import pytz
import asyncio
import time
import math
import random
import functools
import yaml
import certifi
import os
from collections import Counter
import openpyxl
from openpyxl.styles import Alignment

import discord
from discord import Option, OptionChoice
from discord.ext import commands, tasks

from utility.config import config
from cogs.function_in import function_in
from cogs.monster import Monster
from cogs.function_in_in import function_in_in
from cogs.lottery import Lottery
from cogs.skill import Skill
from cogs.quest import Quest_system
from cogs.event import Event

class System(discord.Cog, name="主系統"):
    def __init__(self, bot):
        self.bot: discord.Bot = bot
    
    @discord.Cog.listener()
    async def on_ready(self):
        search = await function_in.sql_findall("rpg_players", "players")
        for user in search:
            await function_in.sql_update("rpg_players", "players", "actioning", "None", "user_id", user[0])
            await function_in.sql_update("rpg_players", "players", "action", 0, "user_id", user[0])
        self.bot.log.info(f'已完成初始化')
    
    @discord.Cog.listener()
    async def on_application_command(self, interaction: discord.Interaction):
        if interaction.guild is None:
            return
        search = await function_in.sql_search("rpg_system", "last_channel", ["guild_id"], [interaction.guild.id])
        if not search:
            await function_in.sql_insert("rpg_system", "last_channel", ["guild_id", "channel_id"], [interaction.guild.id, interaction.channel.id])
        else:
            channel = interaction.guild.get_channel(search[1])
            channel1 = interaction.channel
            if channel != channel1:
                await function_in.sql_update("rpg_system", "last_channel", "channel_id", interaction.channel.id, "guild_id", interaction.guild.id)
    
    @discord.Cog.listener()
    async def on_message(self, msg: discord.Message):
        if msg.channel.id in {1382636864602767460, 1382637390832730173, 1382638422971383861, 1382638616857022635}:
            await msg.publish()
            self.bot.log.info(f'已自動發布訊息\n連結: {msg.jump_url}')

    @commands.slash_command(name="註冊", description="註冊帳號")
    async def 註冊(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        player = interaction.user
        search = await function_in.sql_search("rpg_system", "banlist", ["user_id"], [player.id])
        if search:
            await interaction.followup.send(f'你當前已經被停權了!\n原因: {search[1]}')
            return
        search = await function_in.sql_search("rpg_players", "players", ["user_id"], [player.id])
        if search:
            await interaction.followup.send('你已經註冊過了!')
            return
        embed = discord.Embed(title=f'{player.name} 的註冊選單', color=0x6A6AFF)
        embed.add_field(name="歡迎加入幻境之旅 RPG!", value="\u200b", inline=False)
        embed.add_field(name="請選擇你的職業", value="\u200b", inline=False)
        await interaction.followup.send(embed=embed, view=System.register(self.bot, interaction, player))

    @commands.slash_command(name="復活", description="復活自己")
    async def 復活(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp > 0:
            await interaction.followup.send('你目前並沒有死亡!')
            return
        embed = discord.Embed(title=f'{user.name} 請選擇你的復活方式...', color=0xbe77ff)
        if players_level <= 10:
            embed.add_field(name=f"👼 新手復活", value="復活後不會損失任何經驗(10等及以下可使用)", inline=True)
        embed.add_field(name=f"<:exp:1078583848381710346> 普通復活", value="復活後會損失當前等級滿級所需經驗之30%", inline=True)
        embed.add_field(name=f"<:coin:1078582446091665438> 晶幣復活", value="復活後損失當前等級滿級所需經驗之15%(需要消耗3000晶幣)", inline=True)
        embed.add_field(name=f"<:magic_stone:1078155095126056971> 神聖復活", value="復活後不會損失任何經驗(需要消耗一顆魔法石)", inline=True)
        embed.add_field(name=f"🌎 世界復活", value="復活後不會損失任何經驗(僅限被世界王殺死時使用)", inline=True)
        await interaction.followup.send(embed=embed, view=self.respawn_menu(interaction, players_level))

    @commands.slash_command(name="交易", description="與別人交易",
        options=[
            discord.Option(
                str,
                name="交易選項",
                description="請選擇要交易晶幣或是物品",
                required=True,
                choices=[
                    OptionChoice(name="晶幣", value="晶幣"),
                    OptionChoice(name="水晶", value="水晶"),
                    OptionChoice(name="物品", value="物品")
                ]
            ),
            discord.Option(
                discord.Member,
                name="玩家",
                description="請選擇一個要交易的玩家",
                required=True
            ),
            discord.Option(
                str,
                name="物品",
                description="當交易選項為物品時請於此輸入要交易的物品",
                required=False
            ),
            discord.Option(
                int,
                name="金額或數量",
                description="當交易選項為晶幣或水晶時請於此輸入要交易的金額; 若是物品則填入交易數量, 不填默認為1",
                required=False
            )
        ]
    )
    @commands.cooldown(1, 300, commands.BucketType.user)
    async def 交易(self, interaction: discord.ApplicationContext, func: str, players: discord.Member, item: str, num: int):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        checkreg = await function_in.checkreg(self, interaction, players.id)
        if not checkreg:
            return
        if user.id == players.id:
            await interaction.followup.send('你無法於自己交易!')
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        
        if func == "晶幣":
            if not num:
                await interaction.followup.send('請於 `金額` 選項輸入欲交易的金額!')
                return
            check = await function_in.check_money(self, user, "money", num)
            if not check:
                await interaction.followup.send('你沒有足夠的晶幣來完成這筆交易!')
                return
            embed = discord.Embed(title=f'{user.name} 請確認是否交易...', color=0x9F35FF)
            embed.add_field(name=f"交易對象", value=f"{players.mention}", inline=False)
            embed.add_field(name=f"交易金額", value=f"{num}晶幣", inline=False)
            gold = num
            fee_gold = round(num * 0.1)
            embed.add_field(name=f"交易對象將能獲得 {gold} 晶幣", value="\u200b", inline=False)
            embed.add_field(name=f"手續費", value=f"{fee_gold} 晶幣(10%)", inline=False)
            embed.add_field(name=f"是否接受?", value="\u200b", inline=False)
            embed.add_field(name=f"請在一分鐘內選擇, 逾時自動取消", value="\u200b", inline=False)
            checkactioning, stat = await function_in.checkactioning(self, user, "交易")
            if not checkactioning:
                await interaction.followup.send(f'你當前正在 {stat} 中, 無法交易!')
                return
            await interaction.followup.send(embed=embed, view=self.trade(interaction, players, func, num))
        if func == "水晶":
            if not num:
                await interaction.followup.send('請於 `金額` 選項輸入欲交易的金額!')
                return
            check = await function_in.check_money(self, user, "diamond", num)
            if not check:
                await interaction.followup.send('你沒有足夠的水晶來完成這筆交易!')
                return
            embed = discord.Embed(title=f'{user.name} 請確認是否交易...', color=0x9F35FF)
            embed.add_field(name=f"交易對象", value=f"{players.mention}", inline=False)
            embed.add_field(name=f"交易金額", value=f"{num} 水晶", inline=False)
            gold = num
            fee_gold = round(num * 0.1)
            embed.add_field(name=f"交易對象將能獲得 {gold} 水晶", value="\u200b", inline=False)
            embed.add_field(name=f"手續費", value=f"{fee_gold} 水晶(10%)", inline=False)
            embed.add_field(name=f"是否接受?", value="\u200b", inline=False)
            embed.add_field(name=f"請在一分鐘內選擇, 逾時自動取消", value="\u200b", inline=False)
            checkactioning, stat = await function_in.checkactioning(self, user, "交易")
            if not checkactioning:
                await interaction.followup.send(f'你當前正在 {stat} 中, 無法交易!')
                return
            await interaction.followup.send(embed=embed, view=self.trade(interaction, players, func, num))
        if func == "物品":
            if not item:
                await interaction.followup.send('請於 `物品` 選項輸入欲交易的物品!')
                return
            if not num:
                num = 1
            data = await function_in.search_for_file(self, item)
            if not data:
                await interaction.followup.send(f"`{item}` 不存在於資料庫! 請聯繫GM處理!")
                return
            check, numa = await function_in.check_item(self, user.id, item, num)
            if not check:
                if numa <= 0:
                    await interaction.followup.send(f"你沒有 `{item}` !")
                    return
                await interaction.followup.send(f'你只有 {numa} 個 `{item}` !')
            gold = num*10
            check = await function_in.check_money(self, user, "money", gold)
            if not check:
                await interaction.followup.send(f"交易物品需要收取{gold}晶幣手續費!")
                return
                
            embed = discord.Embed(title=f'{user.name} 請確認是否交易...', color=0x9F35FF)
            embed.add_field(name=f"交易對象", value=f"{players.mention}", inline=False)
            embed.add_field(name=f"交易物品", value=f"{item}", inline=False)
            embed.add_field(name=f"交易件數", value=f"{num}", inline=False)
            embed.add_field(name=f"手續費", value=f"{gold}晶幣", inline=False)
            embed.add_field(name=f"是否接受?", value="\u200b", inline=False)
            embed.add_field(name=f"請在一分鐘內選擇, 逾時自動取消", value="\u200b", inline=False)
            checkactioning, stat = await function_in.checkactioning(self, user, "交易")
            if not checkactioning:
                await interaction.followup.send(f'你當前正在 {stat} 中, 無法交易!')
                return
            await interaction.followup.send(embed=embed, view=self.trade(interaction, players, func, item, num))
            
    @交易.error
    async def 交易_error(self, interaction: discord.ApplicationContext, error: Exception):
        if error.retry_after is not None:
            time = await function_in_in.time_calculate(int(error.retry_after))
            await interaction.response.send_message(f'該指令冷卻中! 你可以在 {time} 後再次使用.', ephemeral=True)
            return

    @commands.slash_command(name="傳送", description="切換至其他地圖",
        options=[
            discord.Option(
                str,
                name="地圖",
                description="選擇一張地圖",
                required=True,
                choices=[
                    OptionChoice(name="Lv1-10翠葉林地", value="翠葉林地"),
                    OptionChoice(name="Lv11-20無盡山脊", value="無盡山脊"),
                    OptionChoice(name="Lv21-30極寒之地", value="極寒之地"),
                    OptionChoice(name="Lv31-40熔岩深谷", value="熔岩深谷"),
                    OptionChoice(name="Lv41-50矮人礦山", value="矮人礦山"),
                    OptionChoice(name="Lv51-60幽暗迷宮", value="幽暗迷宮")
                ],
            )
        ]
    )
    async def 傳送(self, interaction: discord.ApplicationContext, map: str):
        await interaction.defer()
        user = interaction.user
        checkaction = await function_in.checkaction(self, interaction, user.id, config.cd_傳送)
        if not checkaction:
            return
        checkactioning, stat = await function_in.checkactioning(self, user)
        if not checkactioning:
            await interaction.followup.send(f'你當前正在 {stat} 中, 無法傳送!')
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        if f"{map}" == f"{players_map}":
            await interaction.followup.send(f'你傳送後, 發現自己原本就在 {map} 裡面了!')
            return
        await function_in.sql_update("rpg_players", "players", "map", map, "user_id", user.id)
        await interaction.followup.send(f'你成功傳送到 `{map}` !')

    @commands.slash_command(name="背包", description="查看你的背包")
    @commands.cooldown(1, 30, commands.BucketType.user)
    async def 背包(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        workbook = openpyxl.Workbook()
        alignment = Alignment(horizontal='center', vertical='center')
        sheet1 = workbook.active
        sheet2 = workbook.create_sheet(title='材料類')
        sheet3 = workbook.create_sheet(title='技能書類')
        sheet4 = workbook.create_sheet(title='裝備類')
        sheet5 = workbook.create_sheet(title='武器類')
        sheet6 = workbook.create_sheet(title='飾品類')
        sheet7 = workbook.create_sheet(title='寵物類')
        sheet8 = workbook.create_sheet(title='卡牌類')
        sheet9 = workbook.create_sheet(title='料理類')
        sheet10 = workbook.create_sheet(title='生成時間')
        sheet1.title = '道具類'
        sheets = workbook.sheetnames
        search = await function_in.sql_findall("rpg_backpack", f"{user.id}")
        sheet1['A1'] = '道具名稱'
        sheet1['B1'] = '數量'
        sheet2['A1'] = '道具名稱'
        sheet2['B1'] = '數量'
        sheet3['A1'] = '道具名稱'
        sheet3['B1'] = '數量'
        sheet4['A1'] = '道具名稱'
        sheet4['B1'] = '數量'
        sheet5['A1'] = '道具名稱'
        sheet5['B1'] = '數量'
        sheet6['A1'] = '道具名稱'
        sheet6['B1'] = '數量'
        sheet7['A1'] = '道具名稱'
        sheet7['B1'] = '數量'
        sheet8['A1'] = '道具名稱'
        sheet8['B1'] = '數量'
        sheet9['A1'] = '道具名稱'
        sheet9['B1'] = '數量'
        sheet10['A1'] = '本工作表生成時間'
        
        msg1 = ""
        msg2 = ""
        msg3 = ""
        msg4 = ""
        msg5 = ""
        msg6 = ""
        msg7 = ""
        msg8 = ""
        msg9 = ""
        a = 0
        b = 0
        c = 0
        d = 0
        e = 0
        f = 0
        g = 0
        h = 0
        i = 0
        for item_info in search:
            item_type = item_info[1]
            name = item_info[0]
            num = item_info[2]
            if item_type == "道具":
                if num > 0:
                    if a < 1:
                        msg1 += f"{name}: {num}個"
                    else:
                        msg1 += f" | {name}: {num}個"
                    a += 1
                    sheet1[f'A{a+1}'] = f'{name}'
                    sheet1[f'B{a+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "材料":
                if num > 0:
                    if b < 1:
                        msg2 += f"{name}: {num}個"
                    else:
                        msg2 += f" | {name}: {num}個"
                    b += 1
                    sheet2[f'A{b+1}'] = f'{name}'
                    sheet2[f'B{b+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "技能書":
                if num > 0:
                    if c < 1:
                        msg3 += f"{name}: {num}個"
                    else:
                        msg3 += f" | {name}: {num}個"
                    c += 1
                    sheet3[f'A{c+1}'] = f'{name}'
                    sheet3[f'B{c+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "裝備":
                if num > 0:
                    if d < 1:
                        msg4 += f"{name}: {num}個"
                    else:
                        msg4 += f" | {name}: {num}個"
                    d += 1
                    sheet4[f'A{d+1}'] = f'{name}'
                    sheet4[f'B{d+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "武器":
                if num > 0:
                    if e < 1:
                        msg5 += f"{name}: {num}個"
                    else:
                        msg5 += f" | {name}: {num}個"
                    e += 1
                    sheet5[f'A{e+1}'] = f'{name}'
                    sheet5[f'B{e+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "飾品":
                if num > 0:
                    if f < 1:
                        msg6 += f"{name}: {num}個"
                    else:
                        msg6 += f" | {name}: {num}個"
                    f += 1
                    sheet6[f'A{f+1}'] = f'{name}'
                    sheet6[f'B{f+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "寵物":
                if num > 0:
                    if g < 1:
                        msg7 += f"{name}: {num}個"
                    else:
                        msg7 += f" | {name}: {num}個"
                    g += 1
                    sheet7[f'A{g+1}'] = f'{name}'
                    sheet7[f'B{g+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "卡牌":
                if num > 0:
                    if h < 1:
                        msg8 += f"{name}: {num}個"
                    else:
                        msg8 += f" | {name}: {num}個"
                    h += 1
                    sheet8[f'A{h+1}'] = f'{name}'
                    sheet8[f'B{h+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
            if item_type == "料理":
                if num > 0:
                    if i < 1:
                        msg9 += f"{name}: {num}個"
                    else:
                        msg9 += f" | {name}: {num}個"
                    i += 1
                    sheet9[f'A{i+1}'] = f'{name}'
                    sheet9[f'B{i+1}'] = num
                else:
                    await function_in.sql_delete("rpg_backpack", f"{user.id}", "name", name)
        if msg1 == "":
            msg1 = "無"
            sheet1['A2'] = '無'
            sheet1['B2'] = 'X'
        else:
            if len(msg1) > 2000:
                msg1 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg2 == "":
            msg2 = "無"
            sheet2['A2'] = '無'
            sheet2['B2'] = 'X'
        else:
            if len(msg2) > 2000:
                msg2 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg3 == "":
            msg3 = "無"
            sheet3['A2'] = '無'
            sheet3['B2'] = 'X'
        else:
            if len(msg3) > 2000:
                msg3 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg4 == "":
            msg4 = "無"
            sheet4['A2'] = '無'
            sheet4['B2'] = 'X'
        else:
            if len(msg4) > 2000:
                msg4 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg5 == "":
            msg5 = "無"
            sheet5['A2'] = '無'
            sheet5['B2'] = 'X'
        else:
            if len(msg5) > 2000:
                msg5 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg6 == "":
            msg6 = "無"
            sheet6['A2'] = '無'
            sheet6['B2'] = 'X'
        else:
            if len(msg6) > 2000:
                msg6 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg7 == "":
            msg7 = "無"
            sheet7['A2'] = '無'
            sheet7['B2'] = 'X'
        else:
            if len(msg7) > 2000:
                msg7 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg8 == "":
            msg8 = "無"
            sheet8['A2'] = '無'
            sheet8['B2'] = 'X'
        else:
            if len(msg8) > 2000:
                msg8 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        if msg9 == "":
            msg9 = "無"
            sheet9['A2'] = '無'
            sheet9['B2'] = 'X'
        else:
            if len(msg9) > 2000:
                msg9 = "由於該類別超過2000字, 該類別無法顯示.\n請查看excel"
        now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime("%Y年%m月%d日-%H:%M:%S")
        sheet10['B1'] = f"{now_time}"
        for sheet_name in sheets:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = alignment
            sheet.column_dimensions['A'].width = 50
            sheet.column_dimensions['B'].width = 10
            if sheet.title == "生成時間":
                sheet.column_dimensions['A'].width = 30
                sheet.column_dimensions['B'].width = 100
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        folder_path = os.path.join(base_path, "excel_cache")
        save_filename = f"{user.id}_backpack.xlsx"
        save_path = os.path.join(folder_path, save_filename)
        workbook.save(save_path)
        workbook.close()
        file = discord.File(save_path)
        await interaction.followup.send('已發送至私聊')
        msg = await user.send('你的背包:')
        await msg.reply(f"道具類\n```{msg1}```")
        await msg.reply(f"材料類\n```{msg2}```")
        await msg.reply(f"技能書類\n```{msg3}```")
        await msg.reply(f"裝備類\n```{msg4}```")
        await msg.reply(f"武器類\n```{msg5}```")
        await msg.reply(f"飾品類\n```{msg6}```")
        await msg.reply(f"寵物類\n```{msg7}```")
        await msg.reply(f"卡牌類\n```{msg8}```")
        await msg.reply(f"料理類\n```{msg9}```")
        await msg.reply(f"背包完整Excel檔", file=file)
        await msg.reply(f"本背包生成時間: {now_time}")
        os.remove(save_path)


    @背包.error
    async def 背包_error(self, interaction: discord.ApplicationContext, error: Exception):
        if error.retry_after is not None:
            time = await function_in_in.time_calculate(int(error.retry_after))
            await interaction.response.send_message(f'該指令冷卻中! 你可以在 {time} 後再次使用.', ephemeral=True)
            return

    @commands.slash_command(name="使用", description="使用道具",
        options=[
            discord.Option(
                str,
                name="道具名稱",
                description="輸入你想使用的道具名稱",
                required=True
            ),
            discord.Option(
                int,
                name="使用數量",
                description="輸入你想使用的道具數量, 不填則默認為1, 最多一次可使用10個",
                required=False
            )
        ]
    )
    async def 使用(self, interaction: discord.ApplicationContext, name: str, num: int):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        if not num:
            num = 1
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        data, floder_name, floder_name1, c = await function_in.search_for_file(self, name, False)
        if f"{floder_name1}" != "道具" and f"{floder_name1}" != "料理" and f"{floder_name1}" != "技能書" or data is None:
            await interaction.followup.send(f"{name} 不在資料庫內或不是道具/料理/技能書無法使用")
            return False
        if players_level < data[f"{name}"]["等級需求"]:
            await interaction.followup.send(f"你的等級不足以使用 {name} !")
            return
        if "本道具無法直接使用, 需要搭配相關功能" in f"{data[f'{name}']['道具介紹']}":
            await interaction.followup.send(f'{name} 無法直接使用!')
            return
        checkactioning, stat = await function_in.checkactioning(self, user)
        if not checkactioning:
            await interaction.followup.send(f'你當前正在 {stat} 中, 無法使用!')
            return
        if num <= 0:
            await interaction.followup.send('使用數量不得小於0個!')
            return
        if num > 10:
            await interaction.followup.send('使用數量不得大於10個!')
            return
        checknum, numa = await function_in.check_item(self, user.id, name, num)
        if not checknum:
            await interaction.followup.send(f'你沒有 {num} 個 {name}! 你只有 {numa}個')
            return
        if not "使用本道具可無視冷卻時間且不會產生冷卻時間" in f"{data[f'{name}']['道具介紹']}":
            checkaction = await function_in.checkaction(self, interaction, user.id, config.cd_使用)
            if not checkaction:
                return
        if not data:
            await interaction.followup.send('系統發生錯誤! 若使用該物品持續發生錯誤, 請嘗試一次僅使用1個!')
            return
        await function_in.remove_item(self, user.id, name, num)
        quest=False
        embed = discord.Embed(title=f'你成功使用了 {num} 個 `{name}`', color=0x0000c6)
        for i in range(num):
            if "給予道具" in data.get(name, {}):
                for attname, value in data.get(name).get("給予道具", {}).items():
                    await function_in.give_item(self, user.id, attname, value)
                    embed.add_field(name=f"你獲得了 {value} 個 {attname}!", value=f"\u200b", inline=False)
            if "給予裝備" in data.get(name, {}):
                for attname, value in data.get(name).get("給予裝備", {}).items():
                    await function_in.give_item(self, user.id, attname, value)
                    embed.add_field(name=f"你獲得了 {value} 個 {attname}!", value=f"\u200b", inline=False)
            if "給予武器" in data.get(name, {}):
                for attname, value in data.get(name).get("給予武器", {}).items():
                    await function_in.give_item(self, user.id, attname, value)
                    embed.add_field(name=f"你獲得了 {value} 個 {attname}!", value=f"\u200b", inline=False)
            if "給予飾品" in data.get(name, {}):
                for attname, value in data.get(name).get("給予飾品", {}).items():
                    await function_in.give_item(self, user.id, attname, value)
                    embed.add_field(name=f"你獲得了 {value} 個 {attname}!", value=f"\u200b", inline=False)
            if "學會技能" in data.get(name, {}):
                for attname, value in data.get(name).get("學會技能", {}).items():
                    if players_skill_point <= 0:
                        embed.add_field(name=f"你沒有天賦點了, 無法學習 {attname} 技能!", value=f"\u200b", inline=False)
                        await function_in.give_item(self, user.id, name)
                        continue
                    embed.add_field(name=f"技能書燒了起來, 一股黑煙竄進了你的身體", value=f"\u200b", inline=False)
                    dataa, class_name, a, b = await function_in.search_for_file(self, attname, False)
                    if not dataa:
                        embed.add_field(name=f"技能 {attname} 不存在於資料庫! 請聯繫GM處理!", value=f"\u200b", inline=False)
                        continue
                    if "本技能全職業皆可學習" not in dataa["技能介紹"]:
                        if class_name != players_class:
                            embed.add_field(name=f"你無法學會 {attname} 技能! 你的職業為 {players_class}! 該技能需要 {class_name} 職業才能學習!", value=f"\u200b", inline=False)
                            continue
                    search = await function_in.sql_search("rpg_skills", f"{user.id}", ["skill"], [attname])
                    if search:
                        embed.add_field(name=f"你已經學會了 {attname} 技能, 無法再次學習!", value=f"\u200b", inline=False)
                        continue
                    await function_in.sql_insert("rpg_skills", f"{user.id}", ["skill", "level", "exp"], [attname, 1, 0])
                    players_skill_point-=1
                    await function_in.sql_update("rpg_players", "players", "skill_point", players_skill_point, "user_id", user.id)
                    embed.add_field(name=f"你成功學會了 {attname} 技能!", value=f"\u200b", inline=False)
            if "增加屬性" in data.get(name, {}):
                for attname, value in data.get(name).get("增加屬性", {}).items():
                    if "回復" in attname:
                        if attname == "血量回復值":
                            if value == "回滿":
                                embed.add_field(name=f"你的血量回滿了!", value=f"\u200b", inline=False)
                                await function_in.heal(self, user.id, "hp", "max")
                                continue
                            a, b = await function_in.heal(self, user.id, "hp", value)
                            if a == "Full":
                                embed.add_field(name=f"你喝完藥水後, 發現血量本來就是滿的, 藥力流失了...", value=f"\u200b", inline=False)
                            else:
                                if b == "Full":
                                    embed.add_field(name=f"恢復了 {a} HP! ({a-value})", value=f"\u200b", inline=False)
                                else:
                                    embed.add_field(name=f"恢復了 {a} HP!", value=f"\u200b", inline=False)
                        elif attname == "魔力回復值":
                            if value == "回滿":
                                embed.add_field(name=f"你的魔力回滿了!", value=f"\u200b", inline=False)
                                await function_in.heal(self, user.id, "mana", "max")
                                continue
                            a, b = await function_in.heal(self, user.id, "mana", value)
                            if a == "Full":
                                embed.add_field(name=f"你喝完藥水後, 發現魔力本來就是滿的, 藥力流失了...", value=f"\u200b", inline=False)
                            else:
                                if b == "Full":
                                    embed.add_field(name=f"恢復了 {a} MP! ({a-value})", value=f"\u200b", inline=False)
                                else:
                                    embed.add_field(name=f"恢復了 {a} MP!", value=f"\u200b", inline=False)
                        elif attname == "血量回復百分比":
                            hps = int(players_max_hp * (value*0.01))
                            a, b = await function_in.heal(self, user.id, "hp", hps)
                            if a == "Full":
                                embed.add_field(name=f"你喝完藥水後, 發現血量本來就是滿的, 藥力流失了...", value=f"\u200b", inline=False)
                            else:
                                if b == "Full":
                                    embed.add_field(name=f"恢復了 {a} HP! ({a-value})", value=f"\u200b", inline=False)
                                else:
                                    embed.add_field(name=f"恢復了 {a} HP!", value=f"\u200b", inline=False)
                        elif attname == "魔力回復百分比":
                            manas = int(players_max_mana * (value*0.01))
                            a, b = await function_in.heal(self, user.id, "mana", manas)
                            if a == "Full":
                                embed.add_field(name=f"你喝完藥水後, 發現魔力本來就是滿的, 藥力流失了...", value=f"\u200b", inline=False)
                            else:
                                if b == "Full":
                                    embed.add_field(name=f"恢復了 {a} MP! ({a-value})", value=f"\u200b", inline=False)
                                else:
                                    embed.add_field(name=f"恢復了 {a} MP!", value=f"\u200b", inline=False)
                    if "行動條冷卻時間" in attname:
                        if value == "歸零":
                            embed.add_field(name=f"你的行動條歸零了! 你感覺到充滿了體力!", value=f"\u200b", inline=False)
                            await function_in.sql_update("rpg_players", "players", "action", 0, "user_id", user.id)
                    if "晶幣" in attname:
                        await function_in.give_money(self, user, "money", value, "使用道具")
                        embed.add_field(name=f"你獲得了 {value} 晶幣!", value=f"\u200b", inline=False)
                        quest=value
                    if "卡片" in attname:
                        card = await function_in.card_packet(self, attname)
                        await function_in.give_item(self, user.id, card)
                        embed.add_field(name=f"你獲得了 {card}!", value=f"\u200b", inline=False)
                    if "對敵人造成傷害" in attname:
                        await function_in.give_item(self, user.id, name)
                        embed.add_field(name="該道具只能裝備於戰鬥道具欄位, 並於戰鬥中透過快捷欄使用!", value=f"\u200b", inline=False)
                    if "獲得經驗" in attname:
                        expc = await function_in.give_exp(self, user.id, value)
                        embed.add_field(name=f"你獲得了 {value} EXP!", value=f"\u200b", inline=False)
                        if expc:
                            embed.add_field(name=expc, value=f"\u200b", inline=False)
                    if "屬性重置" in attname:
                        await function_in.sql_update("rpg_players", "players", "attr_point", players_level, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_str", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_int", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_dex", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_con", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_luk", 0, "user_id", user.id)
                        embed.add_field(name="成功重置所有屬性點!", value=f"\u200b", inline=False)
                    if "屬性點增加" in attname:
                        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
                        await function_in.sql_update("rpg_players", "players", "add_attr_point", players_add_attr_point+value, "user_id", user.id)
                        embed.add_field(name=f"成功獲得 {value} 點屬性點!", value=f"\u200b", inline=False)
                    if "全屬性增加" in attname:
                        search = await function_in.sql_search("rpg_players", "players", ["user_id"], [user.id])
                        players_all_attr_point = search[20]
                        if int(players_level*0.1)*5 < players_all_attr_point:
                            embed.add_field(name=f"你當前已無法在使用更多的 {name}!", value=f"\u200b", inline=False)
                        else:
                            await function_in.sql_update("rpg_players", "players", "all_attr_point", players_all_attr_point+value, "user_id", user.id)
                            embed.add_field(name=f"力量+{value}!\n智慧+{value}!\n敏捷+{value}!\n體質+{value}!\n幸運+{value}!", value=f"\u200b", inline=False)
                    if "任務放棄" in attname:
                        search = await function_in.sql_search("rpg_players", "quest", ["user_id"], [user.id])
                        if not search:
                            embed.add_field(name="你使用了任務放棄證明後才發現\n你根本沒有接任務阿...", value=f"\u200b", inline=False)
                        else:
                            await function_in.sql_delete("rpg_players", "quest", "user_id", user.id)
                            embed.add_field(name="你成功放棄了當前任務!", value=f"\u200b", inline=False)
                    if "水晶" in attname:
                        if value == "0-20":
                            diamond1 = random.randint(0, 20)
                            embed.add_field(name=f"你獲得了 {diamond1} 顆水晶!", value=f"\u200b", inline=False)
                            await function_in.give_money(self, user, "diamond", diamond1, "道具")
                    if "BOSS召喚" in attname:
                        await function_in.sql_update("rpg_players", "players", "boss", True, "user_id", user.id)
                        embed.add_field(name=f"你下次攻擊必定召喚出Boss!", value=f"\u200b", inline=False)
                    if "簽到禮包" in attname:
                        lot_list = {
                            "普通卡包": 60,
                            "晶幣袋(100元)": 40,
                            "晶幣袋(1000元)": 30,
                            "晶幣袋(2000元)": 20,
                            "晶幣袋(5000元)": 10,
                            "小型經驗包": 10,
                            "稀有卡包": 5,
                            "超級好運卷軸": 5,
                            "Boss召喚卷": 5,
                            "魔法石": 5,
                            "詛咒之石": 2,
                            "晶幣袋(10000元)": 1,
                        }
                        for i in range(num):
                            item = await function_in.lot(self, lot_list)
                            data = await function_in.search_for_file(self, item)
                            if not data:
                                item = "簽到禮包"
                            embed.add_field(name=f"你獲得了1個 {item} !", value=f"\u200b", inline=False)
                            await function_in.give_item(self, user.id, item)
                    if "卡牌欄位解鎖" in attname:
                        search = await function_in.sql_search("rpg_equip", f"{user.id}", ["slot"], ["卡牌欄位2"])
                        slot2 = search[1]
                        search = await function_in.sql_search("rpg_equip", f"{user.id}", ["slot"], ["卡牌欄位3"])
                        slot3 = search[1]
                        if "普通" in attname:
                            if slot2 == "未解鎖":
                                await function_in.sql_update("rpg_equip", f"{user.id}", "equip", "無", "slot", "卡牌欄位2")
                                embed.add_field(name=f"你成功解鎖了卡牌欄位2!", value=f"\u200b", inline=False)
                            else:
                                embed.add_field(name=f"你的卡牌欄位2已經解鎖了! 你的 {name} 燒毀了...", value=f"\u200b", inline=False)
                        elif "高級" in attname:
                            if slot2 == "未解鎖":
                                embed.add_field(name=f"你的卡牌欄位2尚未解鎖! {name} 燒毀了...", value=f"\u200b", inline=False)
                            else:
                                if slot3 == "未解鎖":
                                    await function_in.sql_update("rpg_equip", f"{user.id}", "equip", "無", "slot", "卡牌欄位3")
                                    embed.add_field(name=f"你成功解鎖了卡牌欄位3!", value=f"\u200b", inline=False)
                                else:
                                    embed.add_field(name=f"你的卡牌欄位3已經解鎖了! 你的 {name} 燒毀了...", value=f"\u200b", inline=False)
                    if "經驗加倍" in attname:
                        now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
                        timeString = now_time
                        struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
                        time_stamp = int(time.mktime(struct_time)) + 3600
                        if "全服" in attname:
                            search = await function_in.sql_search("rpg_exp", "all", ["user_id"], [user.id])
                            if not search:
                                await function_in.sql_insert("rpg_exp", "all", ["user_id", "time_stamp", "exp"], [user.id, time_stamp, value])
                                embed.add_field(name=f"你成功啟用了{value}倍全服經驗加倍一小時!", value=f"\u200b", inline=False)
                            else:
                                exp_time_stamp = search[1]
                                exp = search[2]
                                if exp == value:
                                    await function_in.sql_update("rpg_exp", "all", "time_stamp", exp_time_stamp+3600, "user_id", user.id)
                                    embed.add_field(name=f"你成功增加了{value}倍全服經驗加倍一小時!", value=f"\u200b", inline=False)
                                else:
                                    embed.add_field(name=f"你的全服經驗加倍倍數不同! 你的 {name} 回到了你的手中", value=f"\u200b", inline=False)
                                    await function_in.give_item(self, user.id, name)
                        else:
                            search = await function_in.sql_search("rpg_exp", "player", ["user_id"], [user.id])
                            if not search:
                                await function_in.sql_insert("rpg_exp", "player", ["user_id", "time_stamp", "exp"], [user.id, time_stamp, value])
                                embed.add_field(name=f"你成功啟用了{value}倍個人經驗加倍一小時!", value=f"\u200b", inline=False)
                            else:
                                exp_time_stamp = search[1]
                                exp = search[2]
                                if exp == value:
                                    await function_in.sql_update("rpg_exp", "player", "time_stamp", exp_time_stamp+3600, "user_id", user.id)
                                    embed.add_field(name=f"你成功增加了{value}倍個人經驗加倍一小時!", value=f"\u200b", inline=False)
                                else:
                                    embed.add_field(name=f"你的個人經驗加倍倍數不同! 你的 {name} 回到了你的手中", value=f"\u200b", inline=False)
                                    await function_in.give_item(self, user.id, name)
                    if "掉落物" in attname:
                        prizes = {
                            "魔法石": 3000,
                            "水晶箱": 2200,
                            "Boss召喚卷": 1800,
                            "屬性增加藥水": 1450,
                            "史詩卡包": 1250,
                            "傳說卡包": 50,
                            "神性之石": 30,
                            "奇異質點": 1,
                            "「古樹之森」副本入場卷": 1000,
                            "「寒冰之地」副本入場卷": 1000,
                            "「黑暗迴廊」副本入場卷": 1000,
                            "「惡夢迷宮」副本入場卷": 1000,
                        }
                        if "冰霜巨龍" in attname:
                            prizes["冰霜巨龍的鱗片"] = 1500
                            prizes["冰霜巨龍的寶箱"] = 1500
                            prizes["冰霜幼龍"] = 1,
                            prizes["初級天賦領悟書"] = 10
                        if "炎獄魔龍" in attname:
                            prizes["炎獄魔龍的鱗片"] = 1500
                            prizes["炎獄魔龍的寶箱"] = 1500
                            prizes["炎獄幼龍"] = 1,
                            prizes["初級天賦領悟書"] = 10
                        if "魅魔女王" in attname:
                            prizes["魅魔女王的緊身衣碎片"] = 1500
                            prizes["魅魔女王的寶箱"] = 1500
                            prizes["魅魔女王的皮鞭"] = 1
                            prizes["中級天賦領悟書"] = 15
                        
                        item = await function_in.lot(self, prizes)
                        await function_in.give_item(self, user.id, item)
                        embed.add_field(name=f"你獲得了 {item} !", value=f"\u200b", inline=False)
                    if "副本" in attname:
                        dungeon = attname.replace("副本", "")
                        search = await function_in.sql_search("rpg_players", "dungeon", ["user_id"], [user.id])
                        if not search:
                            await function_in.sql_insert("rpg_players", "dungeon", ["user_id", "dungeon_1"], [user.id, 1])
                        if dungeon == "古樹之森":
                            a = 1
                        if dungeon == "寒冰之地":
                            a = 2
                        if dungeon == "黑暗迴廊":
                            a = 3
                        if dungeon == "惡夢迷宮":
                            a = 4
                        if dungeon == "夢魘級惡夢迷宮":
                            a = 5
                        search = await function_in.sql_search("rpg_players", "dungeon", ["user_id"], [user.id])
                        await function_in.sql_update("rpg_players", "dungeon", f"dungeon_{a}", search[a]+1, "user_id", user.id)
                        embed.add_field(name=f"你的{dungeon}副本次數+1!", value=f"\u200b", inline=False)
                    if "料理_" in attname:
                        food = attname.replace("料理_", "")
                        check = await function_in.sql_check_table("rpg_food", f"{user.id}")
                        if not check:
                            await function_in.sql_create_table("rpg_food", f"{user.id}", ["food", "time_stamp"], ["VARCHAR(100)", "BIGINT"], "food")
                        search = await function_in.sql_search("rpg_food", f"{user.id}", ["food"], [food])
                        now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
                        timeString = now_time
                        struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
                        time_stamp = int(time.mktime(struct_time))
                        time_stamp = time_stamp + 3600
                        if not search:
                            await function_in.sql_insert("rpg_food", f"{user.id}", ["food", "time_stamp"], [food, time_stamp])
                        else:
                            await function_in.sql_update("rpg_food", f"{user.id}", "time_stamp", time_stamp, "food", food)
                        embed.add_field(name=f"你成功食用了 {food} !", value=f"\u200b", inline=False)
                    if "飽食度回復" in attname:
                        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
                        players_hunger += value
                        if players_hunger > 100:
                            players_hunger = 100
                        await function_in.sql_update("rpg_players", "players", "hunger", players_hunger, "user_id", user.id)
                        embed.add_field(name=f"你回復了 {value} 點飽食度!", value=f"\u200b", inline=False)
                    if "強化層數" in attname:
                        search = await function_in.sql_search("rpg_players", "equip_upgrade_chance", ["user_id"], [user.id])
                        if value == 0:
                            if not search:
                                await function_in.sql_insert("rpg_players", "equip_upgrade_chance", ["user_id", "amount"], [user.id, 0])
                            else:
                                await function_in.sql_update("rpg_players", "equip_upgrade_chance", "amount", 0, "user_id", user.id)
                            await interaction.followup.send(f"你成功將強化層數堆疊歸零!")
                            continue
                        if search:
                            if search[1] > 0:
                                embed.add_field(name="你當前已擁有強化層數!", value=f"\u200b", inline=False)
                                await function_in.give_item(self, user.id, name)
                                continue
                            else:
                                await function_in.sql_update("rpg_players", "equip_upgrade_chance", "amount", value, "user_id", user.id)
                        else:
                            await function_in.sql_insert("rpg_players", "equip_upgrade_chance", ["user_id", "amount"], [user.id, value])
                        embed.add_field(name=f"你的強化層數為 {value} 層!", value=f"\u200b", inline=False)
                    if "轉職" in attname:
                        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
                        yaml_path = os.path.join(base_path, "rpg", "職業", f"{value}.yml")
                        try:
                            with open(yaml_path, "r", encoding="utf-8") as f:
                                data = yaml.safe_load(f)
                        except Exception as e:
                            await interaction.followup.send(f"職業 {value} 不存在! 請聯繫GM處理!")
                            await function_in.give_item(self, user.id, name)
                            continue
                        await function_in.sql_update("rpg_players", "players", "class", value, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "level", 1, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "exp", 1, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_str", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_int", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_dex", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_con", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_luk", 0, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "attr_point", 3, "user_id", user.id)
                        await function_in.sql_update("rpg_players", "players", "skill_point", 1, "user_id", user.id)
                        await function_in.sql_deleteall("rpg_skills", f"{user.id}")
                        await function_in.sql_deleteall("rpg_equip", f"{user.id}")
                        item_type_list = ["武器","頭盔","胸甲","護腿","鞋子","副手","戒指","項鍊","披風","護身符","戰鬥道具欄位1","戰鬥道具欄位2","戰鬥道具欄位3","戰鬥道具欄位4","戰鬥道具欄位5","技能欄位1","技能欄位2","技能欄位3"]
                        for item_type in item_type_list:
                            await function_in.sql_insert("rpg_equip", f"{user.id}", ["slot", "equip"], [item_type, "無"])
                        await function_in.sql_insert("rpg_equip", f"{user.id}", ["slot", "equip"], ["卡牌欄位1", "無"])
                        item_type_list = ["卡牌欄位2","卡牌欄位3"]
                        for item_type in item_type_list:
                            await function_in.sql_insert("rpg_equip", f"{user.id}", ["slot", "equip"], [item_type, "未解鎖"])
                        embed.add_field(name=f"你成功轉職為 {value}!", value=f"\u200b", inline=False)
                    if "清除資料" in attname:
                        await function_in.delete_player(self, user.id, True)
                        embed.add_field(name=f"你的所有資料已被清除!", value=f"\u200b", inline=False)
                    if "給予隨機職業技能書" in attname:
                        book = await function_in.get_skill_book(self, value)
                        await function_in.give_item(self, user.id, book)
                        embed.add_field(name=f"你獲得了 {book}!", value=f"\u200b", inline=False)
                    if "領悟天賦點" in attname:
                        a = random.randint(1, 100)
                        if value > a:
                            players_skill_point+=1
                            await function_in.sql_update("rpg_players", "players", "skill_point", players_skill_point, "user_id", user.id)
                            embed.add_field(name="你成功領悟到天賦點! 天賦點+1!", value=f"\u200b", inline=False)
                        else:
                            embed.add_field(name="你沒有領悟到天賦點...", value=f"\u200b", inline=False)
        msg = await interaction.followup.send(embed=embed)
        if quest:
            await Quest_system.add_quest(self, user, "賺錢", "道具", value, msg)

    @commands.slash_command(name="休息", description="休息一下, 回個血~")
    async def 休息(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        checkaction = await function_in.checkaction(self, interaction, user.id, config.cd_休息)
        if not checkaction:
            return
        checkactioning, stat = await function_in.checkactioning(self, user)
        if not checkactioning:
            await interaction.followup.send(f'你當前正在 {stat} 中, 無法休息!')
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        reg = round(players_max_hp * 0.2)
        if reg < 30:
            reg = 30
        a, b = await function_in.heal(self, user.id, "hp", reg)
        await function_in.remove_hunger(self, user.id)
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if a == "Full":
            await interaction.followup.send(f'你休息了一下後發現, 身體本來就很好, 💖不需要休息💖!\n目前飽食度剩餘 {players_hunger}')
        else:
            await interaction.followup.send(f'你休息了一下, 感覺身體好了一些! 💗你回復了 {a} 點血量💗!\n目前飽食度剩餘 {players_hunger}')

    @commands.slash_command(name="冥想", description="冥想一下, 回個魔~")
    async def 冥想(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        checkaction = await function_in.checkaction(self, interaction, user.id, config.cd_冥想)
        if not checkaction:
            return
        checkactioning, stat = await function_in.checkactioning(self, user)
        if not checkactioning:
            await interaction.followup.send(f'你當前正在 {stat} 中, 無法冥想!')
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        reg = round(players_max_mana * 0.2)
        if reg < 30:
            reg = 30
        a, b = await function_in.heal(self, user.id, "mana", reg)
        await function_in.remove_hunger(self, user.id)
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if a == "Full":
            await interaction.followup.send(f'你冥想了一下後發現, 精神本來就很好, ✨不需要冥想✨!\n目前飽食度剩餘 {players_hunger}')
        else:
            await interaction.followup.send(f'你冥想了一下, 感覺精神好了一些! 💦你回復了 {a} 點魔力💦!\n目前飽食度剩餘 {players_hunger}')

    @commands.slash_command(name="wiki", description="查看裝備、材料、道具",
        options=[
            discord.Option(
                str,
                name="名稱",
                description="輸入要查詢的名稱",
                required=True
            )
        ]
    )
    @commands.cooldown(1, 10, commands.BucketType.user)
    async def wiki(self, interaction: discord.ApplicationContext, name: str):
        await interaction.defer()
        checkreg = await function_in.checkreg(self, interaction, interaction.user.id)
        if not checkreg:
            return
        data, a, b, c = await function_in.search_for_file(self, name, False)
        if not data:
            await interaction.followup.send(f'`{name}` 不存在於資料庫! 請聯繫GM處理!')
            return
        embed = discord.Embed(title=f'{name}', color=0x28FF28)
        if data.get('技能類型'):
            if a == "特殊":
                embed.add_field(name=f"職業限制: 全職業", value=f"\u200b", inline=False)
            else:
                embed.add_field(name=f"職業限制: {a}", value=f"\u200b", inline=False)
            embed.add_field(name=f"可學習等級: {data['技能等級']}", value="\u200b", inline=False)
            embed.add_field(name=f"技能類型: {data['技能類型']}", value="\u200b", inline=False)
            embed.add_field(name=f"等級上限: {data['等級上限']}", value="\u200b", inline=False)
            if data['技能類型'] == "主動":
                embed.add_field(name=f"消耗魔力: {data['消耗MP']}", value="\u200b", inline=False)
                if not data['冷卻時間']:
                    embed.add_field(name=f"冷卻時間: 0", value="\u200b", inline=False)
                else:
                    embed.add_field(name=f"冷卻時間: {data['冷卻時間']}", value="\u200b", inline=False)
            embed.add_field(name="技能介紹:", value=f"```\n{data['技能介紹']}\n```", inline=False)
        else:
            embed.add_field(name=f"物品類型: {data[f'{name}']['裝備類型']}", value=f"\u200b", inline=False)
            if f"{data[f'{name}']['裝備類型']}" == "寵物":
                embed.add_field(name=f"寵物品級: {data[f'{name}']['寵物品級']}", value=f"\u200b", inline=False)
                a=0
                for attname, value in data.get(name).get("寵物屬性", {}).items():
                    a+=1
                if a > 0:
                    embed.add_field(name=f"寵物屬性: ", value=f"\u200b", inline=False)
                    che_hit = False
                    for attname, value in data.get(name).get("寵物屬性", {}).items():
                        if attname == "命中率":
                            embed.add_field(name=f"\u200b        {attname}: {value+20}", value=f"\u200b", inline=False)
                            che_hit = True
                        else:
                            embed.add_field(name=f"\u200b        {attname}: {value}", value=f"\u200b", inline=False)
                    if not che_hit:
                        embed.add_field(name=f"\u200b        命中率: 20", value=f"\u200b", inline=False)

            elif f"{data[f'{name}']['裝備類型']}" == "勳章":
                embed.add_field(name=f"增加屬性: ", value=f"\u200b", inline=False)
                for attname, value in data.get(name).get("增加屬性", {}).items():
                    embed.add_field(name=f"\u200b        {attname}: {value}", value=f"\u200b", inline=False)

            elif f"{data[f'{name}']['裝備類型']}" == "料理":
                embed.add_field(name=f"料理等級: {data[f'{name}']['料理等級']}", value=f"\u200b", inline=False)
                embed.add_field(name=f"增加屬性: ", value=f"\u200b", inline=False)
                for attname, value in data.get(name).get("增加屬性", {}).items():
                    embed.add_field(name=f"\u200b        {attname}: {value}", value=f"\u200b", inline=False)

            elif f"{data[f'{name}']['裝備類型']}" == "卡牌":
                embed.add_field(name=f"卡牌等級: {data[f'{name}']['卡牌等級']}", value=f"\u200b", inline=False)
                embed.add_field(name=f"等級需求: {data[f'{name}']['等級需求']}", value=f"\u200b", inline=False)
                a = str("全職業" if not '職業限制' in data[f'{name}'] else data[f'{name}']['職業限制'])
                embed.add_field(name=f"職業限制: {a}", value=f"\u200b", inline=False)
                try:
                    embed.add_field(name=f"增加屬性: ", value=f"\u200b", inline=False)
                    for attname, value in data.get(name).get("增加屬性", {}).items():
                        embed.add_field(name=f"\u200b        {attname}: {value}", value=f"\u200b", inline=False)
                except:
                    embed.add_field(name=f"當前該裝備尚未有屬性! ", value=f"\u200b", inline=False)

            else:
                embed.add_field(name=f"等級需求: {data[f'{name}']['等級需求']}", value=f"\u200b", inline=False)
                a = str("全職業" if not '職業限制' in data[f'{name}'] else data[f'{name}']['職業限制'])
                embed.add_field(name=f"職業限制: {a}", value=f"\u200b", inline=False)
                a=0
                for attname, value in data.get(name).get("增加屬性", {}).items():
                    a+=1
                if a > 0:
                    try:
                        embed.add_field(name=f"增加屬性: ", value=f"\u200b", inline=False)
                        for attname, value in data.get(name).get("增加屬性", {}).items():
                            embed.add_field(name=f"\u200b        {attname}: {value}", value=f"\u200b", inline=False)
                    except:
                        embed.add_field(name=f"當前該裝備尚未有屬性! ", value=f"\u200b", inline=False)
                a=0
                for attname, value in data.get(name).get("給予道具", {}).items():
                    a+=1
                if a > 0:
                    embed.add_field(name=f"獲得道具: ", value=f"\u200b", inline=False)
                    for attname, value in data.get(name).get("給予道具", {}).items():
                        embed.add_field(name=f"\u200b        {value} 個 {attname}", value=f"\u200b", inline=False)
                a=0
                for attname, value in data.get(name).get("給予裝備", {}).items():
                    a+=1
                if a > 0:
                    embed.add_field(name=f"獲得裝備: ", value=f"\u200b", inline=False)
                    for attname, value in data.get(name).get("給予裝備", {}).items():
                        embed.add_field(name=f"\u200b        {value} 個 {attname}", value=f"\u200b", inline=False)
                a=0
                for attname, value in data.get(name).get("給予武器", {}).items():
                    a+=1
                if a > 0:
                    embed.add_field(name=f"獲得武器: ", value=f"\u200b", inline=False)
                    for attname, value in data.get(name).get("給予武器", {}).items():
                        embed.add_field(name=f"\u200b        {value} 個 {attname}", value=f"\u200b", inline=False)
                a=0
                for attname, value in data.get(name).get("給予飾品", {}).items():
                    a+=1
                if a > 0:
                    embed.add_field(name=f"獲得飾品: ", value=f"\u200b", inline=False)
                    for attname, value in data.get(name).get("給予飾品", {}).items():
                        embed.add_field(name=f"\u200b        {value} 個 {attname}", value=f"\u200b", inline=False)
            if '套裝效果' in data[f'{name}']:
                embed.add_field(name="套裝效果:", value=f"```\n{data[f'{name}']['套裝效果']}\n```", inline=False)
            embed.add_field(name="物品介紹:", value=f"```\n{data[f'{name}']['道具介紹']}\n```", inline=False)
            if '獲取方式' in data[f'{name}']:
                embed.add_field(name="獲取方式:", value=f"```\n{data[f'{name}']['獲取方式']}\n```", inline=False)
            else:
                embed.add_field(name="獲取方式:", value=f"```無```", inline=False)
        await interaction.followup.send(embed=embed)

    @wiki.error
    async def wiki_error(self, interaction: discord.ApplicationContext, error: Exception):
        if error.retry_after is not None:
            time = await function_in_in.time_calculate(int(error.retry_after))
            await interaction.response.send_message(f'該指令冷卻中! 你可以在 {time} 後再次使用.', ephemeral=True)
            return
    
    @commands.slash_command(name="任務", description="任務")
    async def 任務(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        search = await function_in.sql_search("rpg_players", "quest", ["user_id"], [user.id])
        if search:
            qtype = search[1]
            qname = search[2]
            qnum = search[3]
            qnum_1 = search[4]
            qdaily_money = search[5]
            qdaily_exp = search[6]
            qdaily_qp = search[7]
            rewards = ""
            if qdaily_exp > 0:
                a = qdaily_exp
                rewards+=f"{a}經驗值 "
            if qdaily_money > 0:
                a = qdaily_money
                rewards+=f"{a}晶幣 "
            if qdaily_qp > 0:
                a = qdaily_qp
                rewards+=f"{a}點任務點數"
            embed = discord.Embed(title=f'你目前的任務', color=0xB87070)
            embed.add_field(name=f"任務類型: {qtype}任務", value="\u200b", inline=False)
            if qtype == "擊殺":
                embed.add_field(name=f"你需要擊殺{qnum}隻{qname}", value="\u200b", inline=False)
            if qtype == "賺錢":
                embed.add_field(name=f"你需要透過{qtype}方式賺取{qnum}晶幣", value="\u200b", inline=False)
            if qtype == "工作":
                embed.add_field(name=f"你需要{qname}{qnum}次", value="\u200b", inline=False)
            if qtype == "攻略副本":
                embed.add_field(name=f"你需要完成{qnum}次{qname}副本", value="\u200b", inline=False)
            if qtype == "決鬥":
                if qname == "任意":
                    embed.add_field(name=f"你需要與別人決鬥{qnum}次", value="\u200b", inline=False)
                else:
                    embed.add_field(name=f"你需要與別人決鬥並勝利{qnum}次", value="\u200b", inline=False)
            embed.add_field(name=f"任務獎勵:", value=f"{rewards}", inline=False)
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            embed.add_field(name=f"目前任務進度: ({qnum_1}/{qnum})", value="\u200b", inline=False)
        else:
            players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
            quest_info = await Quest_system.get_quest(self, players_level)
            quest_type = quest_info["qtype"]
            quest_name = quest_info["name"]
            quest_num = quest_info["num"]
            quest_daily = quest_info["daily"]
            embed = discord.Embed(title=f'你接到了新任務', color=0xB87070)
            embed.add_field(name=f"任務類型: {quest_type}任務", value="\u200b", inline=False)
            if quest_type == "擊殺":
                embed.add_field(name=f"你需要擊殺{quest_num}隻{quest_name}", value="\u200b", inline=False)
            if quest_type == "賺錢":
                if quest_name == "打怪":
                    embed.add_field(name=f"你需要透過打怪賺取{quest_num}晶幣", value="\u200b", inline=False)
            if quest_type == "工作":
                embed.add_field(name=f"你需要{quest_name}{quest_num}次", value="\u200b", inline=False)
            if quest_type == "攻略副本":
                embed.add_field(name=f"你需要完成{quest_num}次{quest_name}副本", value="\u200b", inline=False)
            if quest_type == "決鬥":
                if quest_name == "任意":
                    embed.add_field(name=f"你需要與別人決鬥{quest_num}次", value="\u200b", inline=False)
                else:
                    embed.add_field(name=f"你需要與別人決鬥並勝利{quest_num}次", value="\u200b", inline=False)
            rewards = ""
            if quest_daily["exp"] > 0:
                a = quest_daily["exp"]
                rewards+=f"{a}經驗值 "
            if quest_daily["money"] > 0:
                a = quest_daily["money"]
                rewards+=f"{a}晶幣 "
            if quest_daily["qp"] > 0:
                a = quest_daily["qp"]
                rewards+=f"{a}點任務點數"
            embed.add_field(name=f"任務獎勵:", value=f"{rewards}", inline=False)
            await function_in.sql_insert("rpg_players", "quest", ["user_id", "qtype", "qname", "qnum", "qnum_1", "qdaily_money", "qdaily_exp", "qdaily_qp"], [user.id, quest_type, quest_name, quest_num, 0, quest_daily["money"], quest_daily["exp"], quest_daily["qp"]])
        await interaction.followup.send(embed=embed)
    
    @commands.slash_command(name="工作", description="查看或使用工作相關",
        options=[
            discord.Option(
                str,
                name="類別",
                description="選擇你想做的事",
                required=True,
                choices=[
                    OptionChoice(name="伐木",value="伐木"),
                    OptionChoice(name="挖礦",value="挖礦"),
                    OptionChoice(name="釣魚",value="釣魚"),
                    OptionChoice(name="種田",value="種田"),
                    OptionChoice(name="狩獵",value="狩獵"),
                    OptionChoice(name="普通採藥",value="普通採藥"),
                    OptionChoice(name="特殊採藥",value="特殊採藥")
                ]
            ),
            discord.Option(
                int,
                name="次數",
                description="請選擇要進行的次數",
                required=True,
                choices=[
                    OptionChoice(name="一次", value=1),
                    OptionChoice(name="五次", value=5),
                    OptionChoice(name="十次", value=10),
                    OptionChoice(name="三十次", value=30),
                    OptionChoice(name="五十次", value=50),
                    OptionChoice(name="一百次", value=100)
                ]
            )
        ]
    )
    async def 工作(self, interaction: discord.ApplicationContext, ltype: str, func: int):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        if ltype == "特殊採藥":
            if func > 30:
                await interaction.followup.send(f'{ltype} 單次最多只能使用30次採集!')
                return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
        timeString = now_time
        struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
        time_stamp = int(time.mktime(struct_time))
        checkaction = await function_in.checkaction(self, interaction, user.id, config.cd_工作*func)
        if not checkaction:
            return
        checkactioning, stat = await function_in.checkactioning(self, user)
        if not checkactioning:
            await interaction.followup.send(f'你當前正在 {stat} 中, 無法進行工作!')
            return
        item, lifemsg, lifemsg1 = await self.工作採集(ltype)
        if func == 1:
            item, lifemsg, lifemsg1 = await self.工作採集(ltype)
            data = await function_in.search_for_file(self, item)
            if not data:
                await interaction.followup.send(f"{item} 不存在於資料庫! 請聯繫GM處理!")
                return
            await function_in.give_item(self, user.id, item)
            await Quest_system.add_quest(self, user, "工作", lifemsg1, func, msg)
            await function_in.remove_hunger(self, user.id)
            players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
            msg = await interaction.followup.send(f"你辛苦的{lifemsg}後, 得到了1個{item}\n目前飽食度剩餘 {players_hunger}")
            return
        msg1 = await interaction.followup.send(f'正在進行大量{ltype}中, 請稍後')
        msg = f"你辛苦的{lifemsg}後, 得到了下列物品:\n"
        itemlist = []
        for i in range(func):
            item, lifemsg, lifemsg1 = await self.工作採集(ltype)
            data = await function_in.search_for_file(self, item)
            if not data:
                msg += f"{item} 不存在於資料庫! 請聯繫GM處理!"
                continue
            await function_in.give_item(self, user.id, item)
            itemlist.append(item)
        counts = Counter(itemlist)
        for key, value in counts.items():
            msg+=f"{key}x {value}\n"
        now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
        timeString = now_time
        struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
        time_stamp1 = int(time.mktime(struct_time))
        if func == 5:
            await function_in.remove_hunger(self, user.id, 3)
        elif func == 10:
            await function_in.remove_hunger(self, user.id, 6)
        elif func == 30:
            await function_in.remove_hunger(self, user.id, 15)
        elif func == 50:
            await function_in.remove_hunger(self, user.id, 20)
        elif func == 100:
            await function_in.remove_hunger(self, user.id, 40)
        use_time = await function_in_in.time_calculate(time_stamp1-time_stamp)
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        msg+=f"\n總共花費了 {use_time}\n目前飽食度剩餘 {players_hunger}"
        await Quest_system.add_quest(self, user, "工作", lifemsg1, func, msg1)
        await msg1.edit(msg)
        chance = {
            "成功": int(func),
            "失敗": int(1000-func)
        }
        chance = await function_in.lot(self, chance)
        if f"{chance}" == "成功":
            await Event.random_event(self, ltype)
    
    async def 工作採集(self, life):
        if life == "伐木":
            lot_list = {
                "破爛的木頭": 80,
                "普通的木頭": 60,
                "稀有的木頭": 30,
                "高級的木頭": 15,
                "超級的木頭": 5,
                "神級的木頭": 1,
            }
            lmsg = "伐下了一棵樹🪓"
            lmsg1 = "砍伐樹木"
        elif life == "挖礦":
            lot_list = {
                "破爛的礦石": 80,
                "普通的礦石": 60,
                "稀有的礦石": 30,
                "高級的礦石": 15,
                "超級的礦石": 5,
                "神級的礦石": 1,
            }
            lmsg = "敲了好幾個礦⛏"
            lmsg1 = "挖掘礦物"
        elif life == "普通採藥":
            lot_list = {
                "普通生命藥草": 70,
                "高級生命藥草": 30,
                "普通魔力藥草": 70,
                "高級魔力藥草": 30,
            }
            lmsg = "採了一些草藥🌿"
            lmsg1 = "採集草藥"
        elif life == "特殊採藥":
            lot_list = {
                "凋零薔薇": 1,
                "荊棘玫瑰": 1,
                "寒冰薄荷": 1,
                "熔岩花": 1,
                "淨化藥草": 1,
                "劇毒棘刺": 1,
            }
            lmsg = "小心翼翼的採了一點特殊的草藥🌿"
            lmsg1 = "採集草藥"
        elif life == "釣魚":
            lot_list = {
                "小鯉魚": 80,
                "金魚": 60,
                "紅魚": 30,
                "鰻魚": 15,
                "鯨魚": 5,
                "鯊魚": 1,
                "鱷魚": 1,
                "龍蝦": 1
            }
            lmsg = "花了很長時間釣魚🎣"
            lmsg1 = "捕獲魚群"
        elif life == "種田":
            lot_list = {
                "麵粉": 1,
                "鹽巴": 1,
                "糖": 1,
                "薑": 1,
            }
            lmsg = "耕種了一塊土地🌾"
            lmsg1 = "種植作物"
        elif life == "狩獵":
            lot_list = {
                "豬肉": 1,
                "牛肉": 1,
                "羊肉": 1,
                "鹿肉": 1,
                "雞肉": 1,
            }
            lmsg = "狩獵了一隻動物🦌"
            lmsg1 = "狩獵動物"
        else:
            return False, False
        
        item = await function_in.lot(self, lot_list)
        return item, lmsg, lmsg1
    
    @commands.slash_command(name="升級", description="升級技能",
        options=[
            discord.Option(
                str,
                name="技能名稱",
                description="輸入要升級的技能名稱",
                required=True
            )
        ]
    )
    async def 升級(self, interaction: discord.ApplicationContext, skill_name: str):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        if players_skill_point <= 0:
            await interaction.followup.send('你當前天賦點為0!')
            return
        base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        data = False
        folders_to_search = ["特殊", "戰士", "弓箭手", "法師", "刺客"]
        for floder_name in folders_to_search:
            yaml_path = os.path.join(base_path, "rpg", "職業", f"{floder_name}.yml")
            if os.path.exists(yaml_path):
                with open(yaml_path, "r", encoding="utf-8") as f:
                    data = yaml.safe_load(f)
                    if data.get(floder_name).get(skill_name):
                        data = data.get(floder_name).get(skill_name)
                        break
        if not data:
            await interaction.followup.send(f"技能 {skill_name} 技能不存在於資料庫! 請聯繫GM處理")
            return
        search = await function_in.sql_search("rpg_skills", f"{user.id}", ["skill"], [skill_name])
        if not search:
            await interaction.followup.send(f"你還沒有學習 {skill_name} 技能!")
            return
        if search[1] >= data["等級上限"]:
            await interaction.followup.send(f"{skill_name} 技能等級已達上限!")
            return
        if players_skill_point < search[1]:
            await interaction.followup.send(f"你的 {skill_name} 技能等級{search[1]}, 需要消耗 {search[1]} 點天賦點才能升等! 你只有 {players_skill_point} 點天賦點!")
            return
        await function_in.sql_update("rpg_skills", f"{user.id}", "level", search[1]+1, "skill", skill_name)
        await function_in.sql_update("rpg_skills", f"{user.id}", "exp", 0, "skill", skill_name)
        await function_in.sql_update("rpg_players", "players", "skill_point", players_skill_point-search[1], "user_id", user.id)
        await interaction.followup.send(f"你成功消耗了 {search[1]} 點天賦點升級了 {skill_name} 技能! 技能等級 {search[1]+1}!")

    @commands.slash_command(name="屬性點", description="屬性加點")
    async def 屬性點(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        if players_hp <= 0:
            await interaction.followup.send('你當前已經死亡, 無法使用本指令')
            return
        if players_attr_point+players_add_attr_point <= 0:
            await interaction.followup.send('你當前屬性點為0!')
            return
        embed = discord.Embed(title=f'{user} 的屬性點加點', color=0x28FF28)
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
        embed.add_field(name="力量:", value=f"{players_str}", inline=False)
        embed.add_field(name="智慧:", value=f"{players_int}", inline=False)
        embed.add_field(name="敏捷:", value=f"{players_dex}", inline=False)
        embed.add_field(name="體質:", value=f"{players_con}", inline=False)
        embed.add_field(name="幸運:", value=f"{players_luk}", inline=False)
        embed.add_field(name="\u200b", value="\u200b", inline=False)
        embed.add_field(name=f"你當前還有 {players_attr_point+players_add_attr_point} 點屬性點", value="\u200b", inline=False)
        await interaction.followup.send(embed=embed, view=System.attr_up(interaction))
    
    @commands.slash_command(name="經驗加倍", description="查看當前經驗加倍")
    async def 經驗加倍(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        embed = discord.Embed(title=f'<:exp:1078583848381710346> 當前經驗加倍資訊', color=0x53FF53)
        add_exp = 0.0
        all_exp_list = await function_in.sql_findall("rpg_exp", "all")
        if all_exp_list:
            for exp_info in all_exp_list:
                add_exp += exp_info[2]
        embed.add_field(name="當前全服經驗加倍倍率:", value=f"{add_exp}倍", inline=False)
        if all_exp_list:
            all_exp_list = await function_in.sql_search("rpg_exp", "all", ["user_id"], [user.id])
            if all_exp_list:
                exp_time_stamp = all_exp_list[1]
                exp = all_exp_list[2]
                now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
                timeString = now_time
                struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
                time_stamp = int(time.mktime(struct_time))
                exp_time = await function_in_in.time_calculate(exp_time_stamp-time_stamp)
                embed.add_field(name="你開啟的全服經驗加倍倍率:", value=f"{exp}倍", inline=False)
                embed.add_field(name="你開啟的全服經驗加倍剩餘時間:", value=f"{exp_time}", inline=False)
        add_exp = 0.0
        player_exp_list = await function_in.sql_search("rpg_exp", "player", ["user_id"], [user.id])
        if player_exp_list:
            add_exp += player_exp_list[2]
            exp_time_stamp = player_exp_list[1]
            now_time = datetime.datetime.now(pytz.timezone("Asia/Taipei")).strftime('%Y-%m-%d %H:%M:%S')
            timeString = now_time
            struct_time = time.strptime(timeString, "%Y-%m-%d %H:%M:%S")
            time_stamp = int(time.mktime(struct_time))
            exp_time = await function_in_in.time_calculate(exp_time_stamp-time_stamp)
        embed.add_field(name="當前個人經驗加倍倍率:", value=f"{add_exp}倍", inline=False)
        if player_exp_list:
            embed.add_field(name="當前個人經驗加倍剩餘時間:", value=f"{exp_time}", inline=False)
        await interaction.followup.send(embed=embed)
    
    @commands.slash_command(name="fix", description="修復資料")
    async def fix(self, interaction: discord.ApplicationContext):
        await interaction.defer()
        user = interaction.user
        await function_in.fixplayer(self, user.id)
        await interaction.followup.send('已修復完您的資料!')
    
    @commands.slash_command(name="垃圾桶", description="丟棄物品",
        options=[
            discord.Option(
                str,
                name="物品",
                description="選擇你要丟棄的物品",
                required=True
            ),
            discord.Option(
                int,
                name="數量",
                description="選擇你要丟棄的數量, 不填默認為1",
                required=False,
                choices=[
                    OptionChoice(name="查看好感度", value=0)
                ]
            )
        ]
    )
    async def 垃圾桶(self, interaction: discord.ApplicationContext, item: str, num: int):
        await interaction.defer()
        user = interaction.user
        checkreg = await function_in.checkreg(self, interaction, user.id)
        if not checkreg:
            return
        if not num:
            num = 1
        data = await function_in.search_for_file(self, item)
        if not data:
            await interaction.followup.send(f"`{item}` 不存在於資料庫! 請聯繫GM處理!")
            return
        checknum, numa = await function_in.check_item(self, user.id, item, num)
        if not checknum:
            await interaction.followup.send(f'你沒有 {num} 個 {item}! 你只有 {numa}個')
            return
        await function_in.remove_item(self, user.id, item, num)
        data, floder_name, floder_name1, item_type1 = await function_in.search_for_file(self, item, False)
        await interaction.followup.send(f'你丟棄了 {num} 個 {item_type1}: {item}!')

    async def respawn(self, member: discord.Member, func) -> int:
        await function_in.heal(self, member.id, "hp", "max")
        await function_in.heal(self, member.id, "mana", "max")
        players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, member.id)
        che = 0
        if players_level >= 10 and players_level < 30:
            che = 1
        if players_level >= 30 and players_level < 60:
            che = 2
        if players_level >= 60 and players_level < 150:
            che = 3
        if func > 0:
            fullexp = int(18 * (players_level * (players_level ** 1.5)))
            expa = int(round(fullexp * (func *0.01)))
            if expa <= 0:
                expa = 1
            players_exp -= expa
            if players_exp < 0:
                players_exp = 0
            await function_in.sql_update("rpg_players", "players", "exp", players_exp, "user_id", member.id)
            return expa
    
    class register(discord.ui.View):
        def __init__(self, bot: discord.Bot, interaction: discord.ApplicationContext, player: discord.Member):
            super().__init__(timeout=30)
            self.interaction = interaction
            self.bot = bot
            self.player = player
            self.button1 = discord.ui.Button(label="戰士", style=discord.ButtonStyle.blurple, custom_id="button1")
            self.button2 = discord.ui.Button(label="弓箭手", style=discord.ButtonStyle.red, custom_id="button2")
            self.button3 = discord.ui.Button(label="法師", style=discord.ButtonStyle.green, custom_id="button3")
            self.button4 = discord.ui.Button(label="刺客", style=discord.ButtonStyle.grey, custom_id="button4")
            self.button1.callback = functools.partial(self.button1_callback, interaction)
            self.button2.callback = functools.partial(self.button2_callback, interaction)
            self.button3.callback = functools.partial(self.button3_callback, interaction)
            self.button4.callback = functools.partial(self.button4_callback, interaction)
            self.add_item(self.button1)
            self.add_item(self.button2)
            self.add_item(self.button3)
            self.add_item(self.button4)

        async def on_timeout(self):
            await super().on_timeout()
            self.disable_all_items()
            await self.interaction.followup.send('選擇職業超時! 請重新選擇!', view=None)

        async def button1_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            await self.class_select(interaction, "戰士")
            self.stop()
        
        async def button2_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            await self.class_select(interaction, "弓箭手")
            self.stop()
        
        async def button3_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            await self.class_select(interaction, "法師")
            self.stop()

        async def button4_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            await self.class_select(interaction, "刺客")
            self.stop()
        
        async def class_select(self, interaction: discord.ApplicationContext, players_class):
            user = self.player
            await function_in.register_player(self, user.id, players_class)
            embed = discord.Embed(title=f'{user.name} 註冊成功!', color=0x28FF28)
            embed.add_field(name=f"你的職業是 {players_class}", value="\u200b", inline=False)
            embed.add_field(name=f"歡迎遊玩幻境之旅 RPG!", value="\u200b", inline=False)
            embed.add_field(name=f"記得使用 /幫助 來查看遊戲玩法喔!", value="\u200b", inline=False)
            embed.add_field(name=f"官方Discord群:", value=f"https://www.rbctw.net/discord", inline=False)
            if players_class == "戰士":
                await function_in.give_item(self, user.id, "心靈武器--熱血")
            elif players_class == "弓箭手":
                await function_in.give_item(self, user.id, "心靈武器--斥候")
            elif players_class == "法師":
                await function_in.give_item(self, user.id, "心靈武器--知識")
            elif players_class == "刺客":
                await function_in.give_item(self, user.id, "心靈武器--暗影")
            await interaction.message.edit(embed=embed, view=None)
            guild = self.bot.get_guild(config.guild)
            if interaction.guild == guild:
                role = guild.get_role(config.player)
                await user.add_roles(role)
            self.stop()
        
        async def interaction_check(self, interaction: discord.ApplicationContext) -> bool:
            if interaction.user != self.interaction.user:
                await interaction.response.send_message('你無法替他人選擇職業!', ephemeral=True)
                return False
            else:
                return True

    class trade(discord.ui.View):
        def __init__(self, interaction: discord.ApplicationContext, player: discord.Member, func: str, num, numa=None):
            super().__init__(timeout=60)
            self.interaction = interaction
            self.player = player
            self.func = func
            if func == "晶幣":
                self.money = num
            elif func == "水晶":
                self.money = num
            elif func == "物品":
                self.item = num
                self.num = numa
            self.button1 = discord.ui.Button(label="確認交易", style=discord.ButtonStyle.blurple, custom_id="button1")
            self.button2 = discord.ui.Button(label="取消交易", style=discord.ButtonStyle.red, custom_id="button2")
            self.button1.callback = functools.partial(self.button1_callback, interaction)
            self.button2.callback = functools.partial(self.button2_callback, interaction)
            self.add_item(self.button1)
            self.add_item(self.button2)

        async def on_timeout(self):
            await super().on_timeout()
            self.disable_all_items()
            if self.interaction.message:
                try:
                    msg = await self.interaction.message.edit(view=self)
                    await msg.reply('交易超時! 交易已自動取消!')
                    await function_in.checkactioning(self, self.interaction.user, "return")
                    self.stop()
                except discord.errors.ApplicationCommandInvokeError:
                    pass
            else:
                await self.interaction.followup.send('交易超時! 交易已自動取消!')
                await function_in.checkactioning(self, self.interaction.user, "return")
                self.stop()

        async def button1_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            if self.func == "晶幣":
                gold = self.money
                gold1 = round(self.money*1.1)
                gold2 = round(self.money*0.1)
                await function_in.remove_money(self, interaction.user, "money", gold1)
                await function_in.give_money(self, self.player, "money", gold, "交易", msg)
                embed = discord.Embed(title=f'{interaction.user.name} 交易成功', color=0x28FF28)
                embed.add_field(name=f"{interaction.user} 付出 {gold1} 晶幣", value=f"\u200b", inline=False)
                embed.add_field(name=f"{self.player} 獲得 {gold} 晶幣", value=f"\u200b", inline=False)
                embed.add_field(name=f"{interaction.user} 支付了 {gold2} 晶幣 手續費", value=f"\u200b", inline=False)
            elif self.func == "水晶":
                gold = self.money
                gold1 = round(self.money*1.1)
                gold2 = round(self.money*0.1)
                await function_in.remove_money(self, interaction.user, "diamond", gold1)
                await function_in.give_money(self, self.player, "diamond", gold, "交易", msg)
                embed = discord.Embed(title=f'{interaction.user.name} 交易成功', color=0x28FF28)
                embed.add_field(name=f"{interaction.user} 付出 {gold1} 水晶", value=f"\u200b", inline=False)
                embed.add_field(name=f"{self.player} 獲得 {gold} 水晶", value=f"\u200b", inline=False)
                embed.add_field(name=f"{interaction.user} 支付了 {gold2} 水晶 手續費", value=f"\u200b", inline=False)
            elif self.func == "物品":
                await function_in.remove_money(self, interaction.user, "money", self.num*10)
                await function_in.remove_item(self, interaction.user.id, self.item, self.num)
                await function_in.give_item(self, self.player.id, self.item, self.num)
                embed = discord.Embed(title=f'{interaction.user.name} 交易成功', color=0x28FF28)
                embed.add_field(name=f"{interaction.user} 減少{self.num}個 `{self.item}`", value=f"\u200b", inline=False)
                embed.add_field(name=f"{interaction.user} 付出{self.num*10}晶幣手續費", value=f"\u200b", inline=False)
                embed.add_field(name=f"{self.player} 獲得{self.num}個 `{self.item}`", value=f"\u200b", inline=False)
            await msg.edit(view=None, embed=embed)
            await function_in.checkactioning(self, interaction.user, "return")
            self.stop()

        async def button2_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            await interaction.defer()
            embed = discord.Embed(title=f'{interaction.user.name} 交易失敗', color=0xFF2D2D)
            embed.add_field(name=f"{interaction.user} 已取消交易!", value=f"\u200b", inline=False)
            await interaction.followup.edit_message(interaction.message.id, embed=embed, view=None)
            await function_in.checkactioning(self, interaction.user, "return")
            self.stop()

        async def interaction_check(self, interaction: discord.ApplicationContext) -> bool:
            if interaction.user != self.interaction.user:
                await interaction.response.send_message('你不能干涉交易!', ephemeral=True)
                return False
            else:
                return True

    class respawn_menu(discord.ui.View):
        def __init__(self, interaction: discord.ApplicationContext, players_level):
            super().__init__(timeout=30)
            self.interaction = interaction
            self.players_level = players_level
            if players_level <= 10:
                self.button0 = discord.ui.Button(emoji="👼", style=discord.ButtonStyle.grey, custom_id="button0")
                self.button0.callback = functools.partial(self.button0_callback, interaction)
                self.add_item(self.button0)            
            self.button1 = discord.ui.Button(emoji="🔮", style=discord.ButtonStyle.grey, custom_id="button1")
            self.button2 = discord.ui.Button(emoji="🪙", style=discord.ButtonStyle.blurple, custom_id="button2")
            self.button3 = discord.ui.Button(emoji="<:magic_stone:1078155095126056971>", style=discord.ButtonStyle.green, custom_id="button3")
            self.button4 = discord.ui.Button(emoji="🌎", style=discord.ButtonStyle.red, custom_id="button4")
            self.button1.callback = functools.partial(self.button1_callback, interaction)
            self.button2.callback = functools.partial(self.button2_callback, interaction)
            self.button3.callback = functools.partial(self.button3_callback, interaction)
            self.button4.callback = functools.partial(self.button4_callback, interaction)
            self.add_item(self.button1)
            self.add_item(self.button2)
            self.add_item(self.button3)
            self.add_item(self.button4)

        async def on_timeout(self):
            await super().on_timeout()
            self.disable_all_items()
            if self.interaction.message:
                try:
                    msg = await self.interaction.message.edit(view=self)
                    await msg.reply('復活超時! 請重新使用指令!')
                    self.stop()
                except discord.errors.ApplicationCommandInvokeError:
                    pass
            else:
                await self.interaction.followup.send('復活超時! 請重新使用指令!')
                self.stop()

        async def button0_callback(self, button, interaction: discord.Interaction):
            self.disable_all_items()
            await interaction.response.defer()
            exp = await System.respawn(self, interaction.user, 0)
            embed = discord.Embed(title=f'{interaction.user.name} 你復活了', color=0x9d9d9d)
            embed.add_field(name=f"你使用了新手復活", value=f"你花費了0經驗值復活", inline=True)
            await interaction.followup.edit_message(interaction.message.id, view=None, embed=embed)
            self.stop()

        async def button1_callback(self, button, interaction: discord.Interaction):
            self.disable_all_items()
            await interaction.response.defer()
            exp = await System.respawn(self, interaction.user, 30)
            embed = discord.Embed(title=f'{interaction.user.name} 你復活了', color=0x9d9d9d)
            embed.add_field(name=f"你使用了普通復活", value=f"你花費了{exp}經驗值復活", inline=True)
            await interaction.followup.edit_message(interaction.message.id, view=None, embed=embed)
            self.stop()

        async def button2_callback(self, button, interaction: discord.Interaction):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = interaction.user
            players_money = await function_in.check_money(self, user, "money", 3000)
            if not players_money:
                embed = discord.Embed(title=f'{user.name} 請選擇你的復活方式...', color=0xbe77ff)
                if self.players_level <= 10:
                    embed.add_field(name=f"👼 新手復活", value="復活後不會損失任何經驗(10等及以下可使用)", inline=True)
                embed.add_field(name=f"<:exp:1078583848381710346> 普通復活", value="復活後會損失當前等級滿級所需經驗之30%", inline=True)
                embed.add_field(name=f"<:coin:1078582446091665438> 晶幣復活", value="復活後損失當前等級滿級所需經驗之15%(需要消耗3000晶幣)", inline=True)
                embed.add_field(name=f"<:magic_stone:1078155095126056971> 神聖復活", value="復活後不會損失任何經驗(需要消耗一顆魔法石)", inline=True)
                embed.add_field(name=f"🌎 世界復活", value="復活後不會損失任何經驗(僅限被世界王殺死時使用)", inline=True)
                embed.add_field(name="\u200b", value="\u200b", inline=True)
                embed.add_field(name=":x: 你沒有足夠的晶幣來復活!", value="\u200b", inline=False)
                await msg.edit(embed=embed, view=System.respawn_menu(interaction, self.players_level))
            else:
                exp = await System.respawn(self, interaction.user, 15)
                embed = discord.Embed(title=f'{user.name} 你復活了', color=0xffe153)
                await function_in.remove_money(self, user, "money", 3000)
                embed.add_field(name=f"你使用了晶幣復活", value=f"你花費了3000晶幣及{exp}經驗值復活", inline=True)
                await msg.edit(embed=embed, view=None)
            self.stop()

        async def button3_callback(self, button, interaction: discord.Interaction):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = interaction.user
            check, num = await function_in.check_item(self, user.id, "魔法石")
            if not check:
                embed = discord.Embed(title=f'{user.name} 請選擇你的復活方式...', color=0xbe77ff)
                if self.players_level <= 10:
                    embed.add_field(name=f"👼 新手復活", value="復活後不會損失任何經驗(10等及以下可使用)", inline=True)
                embed.add_field(name=f"<:exp:1078583848381710346> 普通復活", value="復活後會損失當前等級滿級所需經驗之30%", inline=True)
                embed.add_field(name=f"<:coin:1078582446091665438> 晶幣復活", value="復活後損失當前等級滿級所需經驗之15%(需要消耗3000晶幣)", inline=True)
                embed.add_field(name=f"<:magic_stone:1078155095126056971> 神聖復活", value="復活後不會損失任何經驗(需要消耗一顆魔法石)", inline=True)
                embed.add_field(name=f"🌎 世界復活", value="復活後不會損失任何經驗(僅限被世界王殺死時使用)", inline=True)
                embed.add_field(name="\u200b", value="\u200b", inline=True)
                embed.add_field(name=":x: 你沒有足夠的魔法石來復活!", value="\u200b", inline=False)
                await msg.edit(embed=embed, view=System.respawn_menu(interaction, self.players_level))
            else:
                await function_in.remove_item(self, user.id, "魔法石")
                await System.respawn(self, interaction.user, 0)
                embed = discord.Embed(title=f'{user.name} 你復活了', color=0xbe77ff)
                embed.add_field(name=f"你使用了神聖復活", value=f"你花費了一顆魔法石復活", inline=True)
                await msg.edit(embed=embed, view=None)
            self.stop()

        async def button4_callback(self, button, interaction: discord.Interaction):
            self.disable_all_items()
            await interaction.response.edit_message(view=self)
            msg = interaction.message
            user = interaction.user
            search = await function_in.sql_search("rpg_players", "players", ["user_id"], [user.id])
            wbk = search[21]
            if not wbk:
                embed = discord.Embed(title=f'{user.name} 請選擇你的復活方式...', color=0xbe77ff)
                if self.players_level <= 10:
                    embed.add_field(name=f"👼 新手復活", value="復活後不會損失任何經驗(10等及以下可使用)", inline=True)
                embed.add_field(name=f"<:exp:1078583848381710346> 普通復活", value="復活後會損失當前等級滿級所需經驗之30%", inline=True)
                embed.add_field(name=f"<:coin:1078582446091665438> 晶幣復活", value="復活後損失當前等級滿級所需經驗之15%(需要消耗3000晶幣)", inline=True)
                embed.add_field(name=f"<:magic_stone:1078155095126056971> 神聖復活", value="復活後不會損失任何經驗(需要消耗一顆魔法石)", inline=True)
                embed.add_field(name=f"🌎 世界復活", value="復活後不會損失任何經驗(僅限被世界王殺死時使用)", inline=True)
                embed.add_field(name="\u200b", value="\u200b", inline=True)
                embed.add_field(name=":x: 你並不是被世界王殺死的, 無法接受來自世界的力量!", value="\u200b", inline=False)
                await msg.edit(embed=embed, view=System.respawn_menu(interaction, self.players_level))
            else:
                await System.respawn(self, interaction.user, 0)
                await function_in.sql_update("rpg_players", "players", "world_boss_kill", False, "user_id", user.id)
                embed = discord.Embed(title=f'{user.name} 你復活了', color=0xbe77ff)
                embed.add_field(name=f"你使用了世界復活", value="\u200b", inline=True)
                await msg.edit(embed=embed, view=None)
            self.stop()

        async def interaction_check(self, interaction: discord.Interaction) -> bool:
            if interaction.user != self.interaction.user:
                await interaction.response.send_message('你不能幫別人選擇復活!', ephemeral=True)
                return False
            else:
                return True

    class attr_up(discord.ui.View):
        def __init__(self, interaction: discord.ApplicationContext):
            super().__init__(timeout=20)
            self.interaction = interaction
            self.button1 = discord.ui.Button(emoji="<:str:1087788396447010956>", label="力量+1", style=discord.ButtonStyle.blurple, custom_id="button1")
            self.button1.callback = functools.partial(self.button1_callback, interaction)
            self.add_item(self.button1)
            self.button2 = discord.ui.Button(emoji="<:int:1087789657569382522>", label="智慧+1", style=discord.ButtonStyle.blurple, custom_id="button2")
            self.button2.callback = functools.partial(self.button2_callback, interaction)
            self.add_item(self.button2)
            self.button3 = discord.ui.Button(emoji="<:dex:1087792723609788517>", label="敏捷+1", style=discord.ButtonStyle.blurple, custom_id="button3")
            self.button3.callback = functools.partial(self.button3_callback, interaction)
            self.add_item(self.button3)
            self.button4 = discord.ui.Button(emoji="<:con:1087794018911522997>", label="體質+1", style=discord.ButtonStyle.blurple, custom_id="button4")
            self.button4.callback = functools.partial(self.button4_callback, interaction)
            self.add_item(self.button4)
            self.button5 = discord.ui.Button(emoji="<:luk:1087794455760883784>", label="幸運+1", style=discord.ButtonStyle.blurple, custom_id="button5")
            self.button5.callback = functools.partial(self.button5_callback, interaction)
            self.add_item(self.button5)
            self.button6 = discord.ui.Button(emoji="<:str:1087788396447010956>", label="力量+5", style=discord.ButtonStyle.blurple, custom_id="button6")
            self.button6.callback = functools.partial(self.button6_callback, interaction)
            self.add_item(self.button6)
            self.button7 = discord.ui.Button(emoji="<:int:1087789657569382522>", label="智慧+5", style=discord.ButtonStyle.blurple, custom_id="button7")
            self.button7.callback = functools.partial(self.button7_callback, interaction)
            self.add_item(self.button7)
            self.button8 = discord.ui.Button(emoji="<:dex:1087792723609788517>", label="敏捷+5", style=discord.ButtonStyle.blurple, custom_id="button8")
            self.button8.callback = functools.partial(self.button8_callback, interaction)
            self.add_item(self.button8)
            self.button9 = discord.ui.Button(emoji="<:con:1087794018911522997>", label="體質+5", style=discord.ButtonStyle.blurple, custom_id="button9")
            self.button9.callback = functools.partial(self.button9_callback, interaction)
            self.add_item(self.button9)
            self.button10 = discord.ui.Button(emoji="<:luk:1087794455760883784>", label="幸運+5", style=discord.ButtonStyle.blurple, custom_id="button10")
            self.button10.callback = functools.partial(self.button10_callback, interaction)
            self.add_item(self.button10)
            self.button11 = discord.ui.Button(emoji="<:str:1087788396447010956>", label="力量+10", style=discord.ButtonStyle.blurple, custom_id="button11")
            self.button11.callback = functools.partial(self.button11_callback, interaction)
            self.add_item(self.button11)
            self.button12 = discord.ui.Button(emoji="<:int:1087789657569382522>", label="智慧+10", style=discord.ButtonStyle.blurple, custom_id="button12")
            self.button12.callback = functools.partial(self.button12_callback, interaction)
            self.add_item(self.button12)
            self.button13 = discord.ui.Button(emoji="<:dex:1087792723609788517>", label="敏捷+10", style=discord.ButtonStyle.blurple, custom_id="button13")
            self.button13.callback = functools.partial(self.button13_callback, interaction)
            self.add_item(self.button13)
            self.button14 = discord.ui.Button(emoji="<:con:1087794018911522997>", label="體質+10", style=discord.ButtonStyle.blurple, custom_id="button14")
            self.button14.callback = functools.partial(self.button14_callback, interaction)
            self.add_item(self.button14)
            self.button15 = discord.ui.Button(emoji="<:luk:1087794455760883784>", label="幸運+10", style=discord.ButtonStyle.blurple, custom_id="button15")
            self.button15.callback = functools.partial(self.button15_callback, interaction)
            self.add_item(self.button15)

        async def on_timeout(self):
            await super().on_timeout()
            self.disable_all_items()
            if self.interaction.message:
                try:
                    msg = await self.interaction.message.edit(view=self)
                    await function_in.checkactioning(self, self.interaction.user, "return")
                    await msg.reply('屬性點加點視窗自動關閉!')
                    self.stop()
                except discord.errors.ApplicationCommandInvokeError:
                    pass
            else:
                await self.interaction.followup.send('屬性點加點視窗自動關閉!')
                await function_in.checkactioning(self, self.interaction.user, "return")
                self.stop()

        async def button1_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "力量", "str", 6)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button2_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "智慧", "int", 7)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button3_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "敏捷", "dex", 8)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button4_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "體質", "con", 9)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button5_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "幸運", "luk", 10)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button6_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "力量", "str", 6, 5)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button7_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "智慧", "int", 7, 5)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button8_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "敏捷", "dex", 8, 5)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button9_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "體質", "con", 9, 5)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button10_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "幸運", "luk", 10, 5)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button11_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "力量", "str", 6, 10)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button12_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "智慧", "int", 7, 10)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button13_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "敏捷", "dex", 8, 10)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button14_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "體質", "con", 9, 10)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass

        async def button15_callback(self, button, interaction: discord.ApplicationContext):
            self.disable_all_items()
            try:
                await interaction.response.edit_message(view=self)
                msg = interaction.message
                user = interaction.user
                embed, attnum = await self.add_attr(user, "幸運", "luk", 10, 10)
                if not attnum:
                    await msg.edit(embed=embed)
                else:
                    await msg.edit(embed=embed, view=System.attr_up(interaction))
                self.stop()
            except discord.errors.ApplicationCommandInvokeError:
                self.stop()
                pass
        
        async def add_attr(self, user: discord.Member, attr: str, attr_en, attr_num: int, num: int=1):
            search = await function_in.sql_search("rpg_players", "players", ["user_id"], [user.id])
            players_attr = search[attr_num]
            players_attr_point = search[11]
            players_add_attr_point = search[19]
            if players_attr_point+players_add_attr_point < num:
                embed = await self.embed_craft(user)
                embed.add_field(name=f":x: 你的屬性點不足{num}點, 無法使用 `{attr}+{num}` !", value="\u200b", inline=False)
                return embed, players_attr_point+players_add_attr_point > 0
            else:
                await function_in.sql_update("rpg_players", "players", f"attr_{attr_en}", players_attr+num, "user_id", user.id)
                await function_in.sql_update("rpg_players", "players", "attr_point", players_attr_point-num, "user_id", user.id)
                embed = await self.embed_craft(user)
                return embed, players_attr_point+players_add_attr_point-num > 0

        async def embed_craft(self, user: discord.Member):
            embed = discord.Embed(title=f'{user} 的屬性點加點', color=0x28FF28)
            players_level, players_exp, players_money, players_diamond, players_qp, players_wbp, players_pp, players_hp, players_max_hp, players_mana, players_max_mana, players_dodge, players_hit, players_crit_damage, players_crit_chance, players_AD, players_AP, players_def, players_ndef, players_str, players_int, players_dex, players_con, players_luk, players_attr_point, players_add_attr_point, players_skill_point, players_register_time, players_map, players_class, drop_chance, players_hunger = await function_in.checkattr(self, user.id)
            embed.add_field(name="力量:", value=f"{players_str}", inline=False)
            embed.add_field(name="智慧:", value=f"{players_int}", inline=False)
            embed.add_field(name="敏捷:", value=f"{players_dex}", inline=False)
            embed.add_field(name="體質:", value=f"{players_con}", inline=False)
            embed.add_field(name="幸運:", value=f"{players_luk}", inline=False)
            embed.add_field(name="\u200b", value="\u200b", inline=False)
            embed.add_field(name=f"你當前還有 {players_attr_point+players_add_attr_point} 點屬性點", value="\u200b", inline=False)
            return embed

        async def interaction_check(self, interaction: discord.ApplicationContext) -> bool:
            if interaction.user != self.interaction.user:
                await interaction.response.send_message('你不能加點別人的屬性!', ephemeral=True)
                return False
            else:
                return True

def setup(client: discord.Bot):
    client.add_cog(System(client))
